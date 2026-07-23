from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

import backend.app as app_module
import backend.import_fix_ext  # noqa: F401
import backend.sale_import_ext  # noqa: F401
import backend.vyapar_exact_ext  # noqa: F401


def build_client(tmp_path: Path):
    app_module.DB_PATH = tmp_path / "test-vyapar.db"
    app_module.init_db()
    client = TestClient(app_module.app)
    response = client.post(
        "/api/setup",
        json={
            "business_name": "Kishore Traders",
            "owner_name": "Ayush",
            "username": "admin",
            "password": "1234",
        },
    )
    assert response.status_code == 200, response.text
    return client, {"Authorization": f"Bearer {response.json()['token']}"}


def workbook_bytes(workbook: Workbook) -> bytes:
    handle = BytesIO()
    workbook.save(handle)
    return handle.getvalue()


def test_sale_report_reads_item_details_and_payment_summary(tmp_path):
    client, headers = build_client(tmp_path)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Sale Report"
    summary.append(["Generated on Jul 23, 2026"])
    summary.append([])
    summary.append([])
    summary.append([
        "Date", "Order No", "Invoice No", "Party Name", "GSTIN", "Party Phone No.",
        "Transaction Type", "Total Amount", "Payment Type", "Received/Paid Amount",
        "Balance Due", "Due Date", "Status", "Description",
    ])
    summary.append([
        "26/03/2025", "", "13930", "Jain super market", "", "9999999999",
        "Sale", 6076, "Cash", 5000, 1076, "02/04/2025", "Partial", "",
    ])

    details = workbook.create_sheet("Item Details")
    details.append(["Generated on Jul 23, 2026"])
    details.append([])
    details.append([
        "Date", "Invoice No./Txn No.", "Party Name", "Item Name", "Item Code", "HSN/SAC",
        "Category", "Challan/Order No.", "size", "Quantity", "Unit", "UnitPrice",
        "Discount Percent", "Discount", "Tax Percent", "Tax", "Transaction Type", "Amount",
    ])
    details.append([])
    details.append([
        "26/03/2025", "13930", "Jain super market", "Dana batani new", "DBN50", "", "Grains",
        "", "50 kg", 50, "Kg", 124, 2, 124, 0, 0, "Sale", 6076,
    ])

    response = client.post(
        "/api/import/vyapar",
        headers=headers,
        data={"entity_type": "sales", "dry_run": "false"},
        files={"file": ("SaleReport_test.xlsx", workbook_bytes(workbook), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    assert response.json()["rows_imported"] == 1
    sales = client.get("/api/sales", headers=headers).json()
    assert len(sales) == 1
    sale = client.get(f"/api/sales/{sales[0]['id']}", headers=headers).json()
    assert sale["invoice_no"] == "13930"
    assert sale["party_name"] == "Jain super market"
    assert sale["total"] == 6076
    assert sale["paid"] == 5000
    assert sale["due"] == 1076
    assert sale["items"][0]["item_name"] == "Dana batani new"
    assert sale["items"][0]["size"] == "50 kg"


def test_purchase_report_groups_blank_invoice_rows_by_summary_total(tmp_path):
    client, headers = build_client(tmp_path)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Purchase Report"
    summary.append(["Generated on Jul 23, 2026"])
    summary.append([])
    summary.append([
        "Date", "Order No", "Invoice No", "Party Name", "GSTIN", "Party Phone No.",
        "Transaction Type", "Total Amount", "Payment Type", "Received/Paid Amount",
        "Balance Due", "Due Date", "Status", "Description",
    ])
    summary.append(["07/11/2020", "", "", "Cash Supplier", "", "", "Purchase", 300, "Cash", 100, 200, "", "Partial", ""])
    summary.append(["07/11/2020", "", "", "Cash Supplier", "", "", "Purchase", 700, "Cash", 700, 0, "", "Paid", ""])

    details = workbook.create_sheet("Item Details")
    details.append(["Generated on Jul 23, 2026"])
    details.append([])
    details.append([
        "Date", "Invoice No./Txn No.", "Party Name", "Item Name", "Item Code", "HSN/SAC",
        "Category", "Challan/Order No.", "size", "Quantity", "Unit", "UnitPrice",
        "Discount Percent", "Discount", "Tax Percent", "Tax", "Transaction Type", "Amount",
    ])
    details.append([])
    for name, amount in [("Item A", 100), ("Item B", 200), ("Item C", 300), ("Item D", 400)]:
        details.append(["07/11/2020", "", "Cash Supplier", name, "", "", "", "", "", 1, "pcs", amount, 0, 0, 0, 0, "Purchase", amount])

    response = client.post(
        "/api/import/vyapar",
        headers=headers,
        data={"entity_type": "purchases", "dry_run": "false"},
        files={"file": ("PurchaseReport_test.xlsx", workbook_bytes(workbook), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    assert response.json()["rows_imported"] == 4
    purchases = client.get("/api/purchases", headers=headers).json()
    assert len(purchases) == 2
    assert sorted(row["total"] for row in purchases) == [300, 700]
    assert sorted(row["paid"] for row in purchases) == [100, 700]
    details_by_total = {
        row["total"]: client.get(f"/api/purchases/{row['id']}", headers=headers).json()
        for row in purchases
    }
    assert [line["item_name"] for line in details_by_total[300]["items"]] == ["Item A", "Item B"]
    assert [line["item_name"] for line in details_by_total[700]["items"]] == ["Item C", "Item D"]


def test_party_report_maps_phone_and_receivable_payable_balances(tmp_path):
    client, headers = build_client(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Party Report"
    sheet.append(["Name", "Email", "Phone No.", "Address", "GSTIN", "Receivable Balance", "Payable Balance", "Credit Limit"])
    sheet.append([])
    sheet.append(["Customer A", "", "9826710219", "Indore", "", 36378, 0, ""])
    sheet.append(["Supplier B", "", "9999999999", "Indore", "", 0, 935, ""])
    sheet.append(["", "", "", "", "Total", 36378, 935, ""])

    response = client.post(
        "/api/import/vyapar",
        headers=headers,
        data={"entity_type": "parties", "dry_run": "false"},
        files={"file": ("PartyReport.xlsx", workbook_bytes(workbook), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    assert response.json()["rows_total"] == 2
    parties = client.get("/api/parties", headers=headers).json()
    assert len(parties) == 2
    customer = next(row for row in parties if row["name"] == "Customer A")
    supplier = next(row for row in parties if row["name"] == "Supplier B")
    assert customer["phone"] == "9826710219"
    assert customer["type"] == "customer"
    assert customer["balance"] == 36378
    assert supplier["type"] == "supplier"
    assert supplier["balance"] == 935
