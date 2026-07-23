from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from fastapi import Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import load_workbook

import backend.app as core
import backend.import_fix_ext as old_fix
from backend.app import app, current_user, db, now_iso


PREVIOUS_PARSE = core.parse_upload

EXTRA = {
    "invoice_no": ["sale_no", "sales_no", "transaction_id", "transaction_number", "voucher_no", "voucher_number", "invoice_id", "bill_id"],
    "invoice_date": ["sale_date", "transaction_datetime", "voucher_date"],
    "name": ["party", "customer", "account", "party_account_name"],
    "item_name": ["items", "item_details", "item_service_name", "product_details", "goods_name"],
    "qty": ["item_quantity", "primary_quantity", "sold_quantity", "sale_qty", "actual_quantity"],
    "rate": ["item_price", "unit_rate", "selling_rate", "price_per_unit", "rate_with_tax"],
    "gst_rate": ["tax_percent", "gst_percent", "item_tax_rate"],
    "size": ["item_variant", "variant_name", "package_size", "packing", "pack"],
    "unit": ["item_unit", "quantity_unit", "primary_quantity_unit"],
}
for key, aliases in EXTRA.items():
    current = core.ALIASES.setdefault(key, [])
    for alias in aliases:
        if alias not in current:
            current.append(alias)

LINE_TOTALS = ["line_total", "item_total", "item_amount", "taxable_amount", "net_item_amount", "amount", "total_amount"]


def _sale_report(filename: str) -> bool:
    name = Path(filename or "").name.lower().replace("_", "").replace("-", "")
    return "salereport" in name


def _value(row: dict[str, Any], key: str) -> Any:
    for alias in [key, *core.ALIASES.get(key, [])]:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    for raw_key, value in row.items():
        if value in (None, ""):
            continue
        header = core.normalize_header(raw_key)
        if key == "invoice_no" and (("invoice" in header and "no" in header) or ("bill" in header and "no" in header) or ("transaction" in header and "no" in header)):
            return value
        if key == "invoice_date" and "date" in header:
            return value
        if key == "item_name" and (("item" in header and "name" in header) or ("product" in header and "name" in header)):
            return value
        if key == "qty" and "quantity" in header:
            return value
        if key == "rate" and (("item" in header and "rate" in header) or ("unit" in header and "price" in header)):
            return value
        if key == "name" and (("party" in header and "name" in header) or ("customer" in header and "name" in header)):
            return value
    return ""


def _candidate_score(sheet_name: str, rows: list[dict[str, Any]]) -> int:
    headers = set(rows[0].keys()) if rows else set()
    fake = {header: "x" for header in headers}
    score = 0
    if _value(fake, "invoice_no"): score += 120
    if _value(fake, "invoice_date"): score += 20
    if _value(fake, "item_name"): score += 180
    if _value(fake, "qty"): score += 90
    if _value(fake, "rate") or any(alias in headers for alias in LINE_TOTALS): score += 80
    title = core.normalize_header(sheet_name)
    if any(word in title for word in ("item", "detail", "product", "particular")): score += 250
    if any(word in title for word in ("summary", "overview")): score -= 100
    return score


def _canonical_rows(rows: list[dict[str, Any]], sheet_name: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    carry: dict[str, Any] = {}
    for raw in rows:
        row = dict(raw)
        for key in ("invoice_no", "invoice_date", "name", "item_name", "qty", "rate", "gst_rate", "size", "unit", "sku", "discount", "paid", "payment_mode"):
            value = _value(row, key)
            if value not in (None, ""):
                row[key] = value
        for alias in LINE_TOTALS:
            if row.get(alias) not in (None, ""):
                row["line_total"] = row[alias]
                break
        for key in ("invoice_no", "invoice_date", "name"):
            if row.get(key) not in (None, ""):
                carry[key] = row[key]
            elif key in carry:
                row[key] = carry[key]
        item = str(row.get("item_name") or "").strip()
        if core.normalize_header(item) in {"total", "subtotal", "grand_total", "invoice_total"}:
            continue
        if not item and row.get("qty") in (None, "") and row.get("rate") in (None, "") and row.get("line_total") in (None, ""):
            continue
        row["_source_sheet"] = sheet_name
        output.append(row)
    return output


def _unique_year_invoices(rows: list[dict[str, Any]]) -> None:
    dates: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        invoice = str(row.get("invoice_no") or "").strip()
        if invoice and row.get("invoice_date") not in (None, ""):
            dates[invoice].add(core.normalize_date(row["invoice_date"]))
    repeated = {invoice for invoice, values in dates.items() if len(values) > 1}
    for row in rows:
        invoice = str(row.get("invoice_no") or "").strip()
        if invoice in repeated:
            day = core.normalize_date(row.get("invoice_date")).replace("-", "")
            row["invoice_no"] = f"{invoice}-{day}"


def parse_sale_item_sheet(filename: str, content: bytes) -> list[dict[str, Any]]:
    if not _sale_report(filename) or Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        return PREVIOUS_PARSE(filename, content)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    candidates: list[tuple[int, str, list[dict[str, Any]]]] = []
    for sheet in workbook.worksheets:
        matrix = list(sheet.iter_rows(values_only=True))
        rows = old_fix._rows_from_matrix(matrix)
        candidates.append((_candidate_score(sheet.title, rows), sheet.title, rows))
    if not candidates:
        return PREVIOUS_PARSE(filename, content)
    _, sheet_name, selected = max(candidates, key=lambda entry: (entry[0], len(entry[2])))
    rows = _canonical_rows(selected, sheet_name)
    usable = [row for row in rows if str(row.get("item_name") or "").strip()]
    with_invoice = [row for row in usable if str(row.get("invoice_no") or "").strip()]
    meaningful = [row for row in usable if core.number(row.get("qty"), 0) or core.money(row.get("rate"), 0) or core.money(row.get("line_total"), 0)]
    if not usable:
        raise ValueError(f"{sheet_name} sheet mein item details nahi mili. Wrong blank bills ko import se rok diya gaya.")
    if len(with_invoice) < max(1, len(usable) // 2):
        raise ValueError(f"{sheet_name} sheet mein Invoice/Bill No. nahi mila. Excel file yahan upload karein taaki exact mapping ho sake.")
    if not meaningful:
        raise ValueError(f"{sheet_name} sheet mein Qty/Rate/Amount read nahi hua. Import nahi kiya gaya.")
    _unique_year_invoices(rows)
    return rows


core.parse_upload = parse_sale_item_sheet


def _bad_batches(conn: Any, business_id: int) -> list[dict[str, Any]]:
    batches = conn.execute("SELECT * FROM import_batches WHERE business_id=? AND entity_type='sales' AND COALESCE(status,'')!='rolled_back' ORDER BY id DESC", (business_id,)).fetchall()
    result = []
    for batch in batches:
        sales = conn.execute("SELECT id,total FROM sales WHERE business_id=? AND import_batch_id=?", (business_id, batch["id"])).fetchall()
        if not sales:
            continue
        bad = True
        for sale in sales:
            if abs(float(sale["total"] or 0)) > 0.000001:
                bad = False
                break
            lines = conn.execute("SELECT item_name,rate,line_total FROM sale_items WHERE sale_id=?", (sale["id"],)).fetchall()
            if not lines or any(str(line["item_name"] or "").strip().lower() not in {"imported sale", "imported item"} or abs(float(line["rate"] or 0)) > 0.000001 or abs(float(line["line_total"] or 0)) > 0.000001 for line in lines):
                bad = False
                break
        if bad:
            result.append({"batch_id": int(batch["id"]), "filename": batch["filename"], "transactions": len(sales)})
    return result


def _remove_bad_batch(conn: Any, business_id: int, batch_id: int) -> None:
    sales = conn.execute("SELECT id FROM sales WHERE business_id=? AND import_batch_id=?", (business_id, batch_id)).fetchall()
    for sale in sales:
        lines = conn.execute("SELECT item_id,qty FROM sale_items WHERE sale_id=?", (sale["id"],)).fetchall()
        for line in lines:
            if line["item_id"]:
                conn.execute("UPDATE items SET stock=stock+?,updated_at=? WHERE id=?", (float(line["qty"] or 0), now_iso(), line["item_id"]))
        conn.execute("DELETE FROM stock_movements WHERE business_id=? AND reference_type='sale' AND reference_id=?", (business_id, sale["id"]))
    conn.execute("DELETE FROM sales WHERE business_id=? AND import_batch_id=?", (business_id, batch_id))
    conn.execute("UPDATE import_batches SET status='rolled_back',rows_imported=0,errors_json=? WHERE id=? AND business_id=?", (json.dumps([{"error": "Removed invalid zero-value import"}]), batch_id, business_id))


@app.post("/api/import/cleanup-empty-sales")
def cleanup_empty_sales(execute: bool = Query(False), user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner access required")
    with db() as conn:
        batches = _bad_batches(conn, user["business_id"])
        if execute:
            for batch in batches:
                _remove_bad_batch(conn, user["business_id"], batch["batch_id"])
            conn.execute("DELETE FROM items WHERE business_id=? AND sku LIKE 'IMP-%' AND lower(trim(name)) IN ('imported sale','imported item') AND NOT EXISTS (SELECT 1 FROM sale_items WHERE sale_items.item_id=items.id) AND NOT EXISTS (SELECT 1 FROM purchase_items WHERE purchase_items.item_id=items.id)", (user["business_id"],))
    return {"execute": execute, "batch_count": len(batches), "transaction_count": sum(batch["transactions"] for batch in batches), "batches": batches}


V2_JS = r"""
(() => {
  'use strict';
  loadImportHistory = async function () {
    const [rows, wrongParties, wrongSales] = await Promise.all([
      api('/api/import/batches'),
      api('/api/import/cleanup-party-items', {method:'POST'}),
      api('/api/import/cleanup-empty-sales', {method:'POST'}),
    ]);
    let warning = '';
    if (wrongParties.count) warning += `<div class="info-banner" style="margin-bottom:12px"><b>${wrongParties.count} party names Items mein mile</b><p>Zero-stock aur unused wrong items hi remove honge.</p><button id="cleanup-wrong-party-items" class="btn primary" type="button">Remove Wrong Party Items</button></div>`;
    if (wrongSales.transaction_count) warning += `<div class="info-banner" style="margin-bottom:12px;border-color:#ef476f"><b>${wrongSales.transaction_count} blank ₹0 sales galat import hui</b><p>Sirf “Imported Sale” wale blank bills remove honge.</p><button id="cleanup-empty-sales" class="btn primary" type="button">Remove Blank Wrong Sales</button></div>`;
    const history = rows.map(row => { const error=row.errors?.[0]?.error||''; const removed=row.status==='rolled_back'?' · Removed':''; return `<div class="simple-row"><div><b>${esc(row.entity_type)} · ${esc(row.filename)}</b><small>${niceDate(row.created_at?.slice(0,10))}${removed}${error?` · ${esc(error)}`:''}</small></div><strong>${row.rows_imported}/${row.rows_total}</strong></div>`; }).join('') || emptyText('Abhi koi import nahi hua.');
    $('#import-history').innerHTML = warning + history;
  };
  document.addEventListener('click', async event => {
    const button = event.target.closest('#cleanup-empty-sales');
    if (!button) return;
    if (!confirm('Sirf ₹0 Imported Sale blank bills remove karein?')) return;
    button.disabled = true;
    try { const result=await api('/api/import/cleanup-empty-sales?execute=true',{method:'POST'}); toast(`${result.transaction_count} wrong blank sales removed`); await refreshAll(); await loadImportHistory(); }
    catch(error){ toast(error.message,true); button.disabled=false; }
  });
})();
"""


for route in list(app.router.routes):
    if getattr(route, "path", None) == "/import-fix.js":
        app.router.routes.remove(route)


@app.get("/import-fix.js", include_in_schema=False)
def import_fix_js_v2() -> Response:
    return Response(f"{old_fix.IMPORT_FIX_JS}\n{V2_JS}", media_type="application/javascript", headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


paths = {"/api/import/cleanup-empty-sales", "/import-fix.js"}
routes = [route for route in app.router.routes if getattr(route, "path", None) in paths]
for route in routes:
    app.router.routes.remove(route)
fallback = next((index for index, route in enumerate(app.router.routes) if getattr(route, "path", None) == "/{path:path}"), len(app.router.routes))
app.router.routes[fallback:fallback] = routes
