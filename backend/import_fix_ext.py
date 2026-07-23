from __future__ import annotations

import csv
import io
import re
import unicodedata
from pathlib import Path
from typing import Any

from fastapi import Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, Response
from openpyxl import load_workbook

import backend.app as core
from backend.app import STATIC_DIR, app, current_user, db


ORIGINAL_PARSE_UPLOAD = core.parse_upload
ORIGINAL_PICK = core.pick
ORIGINAL_INSERT_SALE = core.insert_sale
ORIGINAL_INSERT_PURCHASE = core.insert_purchase


# Vyapar uses slightly different headings in different app/export versions.
core.ALIASES.update({
    "name": [
        "party_name", "customer_name", "supplier_name", "account_name",
        "item_name", "product_name", "name",
    ],
    "invoice_no": [
        "invoice_no", "invoice_number", "invoice", "sale_invoice_no",
        "purchase_invoice_no", "transaction_no", "txn_no", "voucher_no",
        "bill_no", "bill_number", "ref_no", "reference_no", "no",
    ],
    "invoice_date": [
        "invoice_date", "transaction_date", "txn_date", "bill_date", "date",
    ],
    "item_name": [
        "item_name", "product_name", "product", "item", "description",
        "particulars", "item_description",
    ],
    "sku": ["item_code", "product_code", "sku", "code", "item_id"],
    "qty": ["qty", "quantity", "item_qty", "sale_quantity", "purchase_quantity"],
    "rate": ["rate", "price", "item_rate", "unit_price", "price_unit", "sale_price", "purchase_price"],
    "gst_rate": ["tax_rate", "gst", "gst_rate", "tax_percentage", "tax"],
    "discount": ["discount", "discount_amount", "total_discount"],
    "paid": ["paid", "paid_amount", "received_amount", "received", "amount_received"],
    "payment_mode": ["payment_mode", "payment_type", "mode", "payment_status"],
})


KNOWN_HEADERS = {
    alias
    for aliases in core.ALIASES.values()
    for alias in aliases
}
KNOWN_HEADERS.update({
    "total", "total_amount", "grand_total", "invoice_amount", "balance",
    "party_balance", "transaction_type", "status",
})


def _unique_headers(values: tuple[Any, ...] | list[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for value in values:
        base = core.normalize_header(value)
        if not base:
            headers.append("")
            continue
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def _header_score(values: tuple[Any, ...] | list[Any]) -> int:
    normalized = {core.normalize_header(value) for value in values if value not in (None, "")}
    exact = len(normalized & KNOWN_HEADERS)
    useful_words = sum(
        1 for value in normalized
        if any(token in value for token in ("invoice", "party", "item", "quantity", "amount", "rate", "date"))
    )
    return exact * 4 + useful_words


def _rows_from_matrix(matrix: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not matrix:
        return []
    scan_limit = min(len(matrix), 50)
    best_index = 0
    best_score = -1
    for index in range(scan_limit):
        score = _header_score(matrix[index])
        if score > best_score:
            best_index = index
            best_score = score
    headers = _unique_headers(matrix[best_index])
    rows: list[dict[str, Any]] = []
    for values_row in matrix[best_index + 1:]:
        row = {
            headers[index]: values_row[index]
            for index in range(min(len(headers), len(values_row)))
            if headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def parse_upload_smart(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        best_rows: list[dict[str, Any]] = []
        best_quality = -1
        for sheet in workbook.worksheets:
            matrix = list(sheet.iter_rows(values_only=True))
            rows = _rows_from_matrix(matrix)
            quality = len(rows)
            if matrix:
                quality += max((_header_score(row) for row in matrix[:50]), default=0) * 1000
            if quality > best_quality:
                best_quality = quality
                best_rows = rows
        return best_rows
    if suffix == ".xls" and core.xlrd is not None:
        workbook = core.xlrd.open_workbook(file_contents=content)
        best_rows: list[dict[str, Any]] = []
        best_quality = -1
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            matrix = [tuple(sheet.cell_value(row, col) for col in range(sheet.ncols)) for row in range(sheet.nrows)]
            rows = _rows_from_matrix(matrix)
            quality = len(rows)
            if matrix:
                quality += max((_header_score(row) for row in matrix[:50]), default=0) * 1000
            if quality > best_quality:
                best_quality = quality
                best_rows = rows
        return best_rows
    if suffix in {".csv", ".txt"}:
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        matrix = [tuple(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]
        return _rows_from_matrix(matrix)
    return ORIGINAL_PARSE_UPLOAD(filename, content)


def pick_compat(row: dict[str, Any], key: str, default: Any = "") -> Any:
    if key == "item_name":
        for alias in core.ALIASES["item_name"]:
            if alias in row and row[alias] not in (None, ""):
                return row[alias]
        return "Imported Sale"
    if key == "rate":
        direct = ORIGINAL_PICK(row, key, None)
        if direct not in (None, ""):
            return direct
        for alias in ("line_total", "item_total", "amount", "total_amount", "invoice_amount", "grand_total", "total"):
            if alias in row and row[alias] not in (None, ""):
                quantity = max(core.number(ORIGINAL_PICK(row, "qty", 1)), 0.0001)
                return core.money(row[alias]) / quantity
        return default
    return ORIGINAL_PICK(row, key, default)


def _copy_payload(payload: Any, **changes: Any) -> Any:
    if hasattr(payload, "model_copy"):
        return payload.model_copy(update=changes)
    if hasattr(payload, "copy"):
        return payload.copy(update=changes)
    for key, value in changes.items():
        setattr(payload, key, value)
    return payload


def _unique_import_invoice(conn: Any, table: str, bid: int, invoice_no: str, invoice_date: str) -> str:
    base = f"{invoice_no}-{invoice_date.replace('-', '')}"
    candidate = base
    sequence = 2
    while conn.execute(
        f"SELECT 1 FROM {table} WHERE business_id=? AND invoice_no=?",
        (bid, candidate),
    ).fetchone():
        candidate = f"{base}-{sequence}"
        sequence += 1
    return candidate


def insert_sale_compat(conn: Any, bid: int, payload: Any, import_batch_id: int | None = None) -> dict[str, Any]:
    if import_batch_id and str(payload.invoice_no or "").strip():
        invoice_no = str(payload.invoice_no).strip()
        invoice_date = core.normalize_date(payload.invoice_date)
        existing = conn.execute(
            "SELECT invoice_date FROM sales WHERE business_id=? AND invoice_no=?",
            (bid, invoice_no),
        ).fetchone()
        if existing and existing["invoice_date"] != invoice_date:
            payload = _copy_payload(
                payload,
                invoice_no=_unique_import_invoice(conn, "sales", bid, invoice_no, invoice_date),
            )
    return ORIGINAL_INSERT_SALE(conn, bid, payload, import_batch_id)


def insert_purchase_compat(conn: Any, bid: int, payload: Any, import_batch_id: int | None = None) -> dict[str, Any]:
    if import_batch_id and str(payload.invoice_no or "").strip():
        invoice_no = str(payload.invoice_no).strip()
        invoice_date = core.normalize_date(payload.invoice_date)
        existing = conn.execute(
            "SELECT invoice_date FROM purchases WHERE business_id=? AND invoice_no=?",
            (bid, invoice_no),
        ).fetchone()
        if existing and existing["invoice_date"] != invoice_date:
            payload = _copy_payload(
                payload,
                invoice_no=_unique_import_invoice(conn, "purchases", bid, invoice_no, invoice_date),
            )
    return ORIGINAL_INSERT_PURCHASE(conn, bid, payload, import_batch_id)


core.parse_upload = parse_upload_smart
core.pick = pick_compat
core.insert_sale = insert_sale_compat
core.insert_purchase = insert_purchase_compat


def _clean_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = re.sub(r"[\u200B-\u200D\u2060\uFEFF]", "", text)
    return re.sub(r"\s+", " ", text).strip().casefold()


def _wrong_party_item_candidates(conn: Any, business_id: int) -> list[dict[str, Any]]:
    parties = {
        _clean_text(row["name"])
        for row in conn.execute("SELECT name FROM parties WHERE business_id=?", (business_id,)).fetchall()
        if _clean_text(row["name"])
    }
    rows = conn.execute(
        """
        SELECT i.id,i.name,i.sku
        FROM items i
        WHERE i.business_id=?
          AND i.sku LIKE 'IMP-%'
          AND ABS(COALESCE(i.stock,0)) < 0.000001
          AND ABS(COALESCE(i.purchase_price,0)) < 0.000001
          AND ABS(COALESCE(i.sale_price,0)) < 0.000001
          AND ABS(COALESCE(i.mrp,0)) < 0.000001
          AND NOT EXISTS (SELECT 1 FROM sale_items si WHERE si.item_id=i.id)
          AND NOT EXISTS (SELECT 1 FROM purchase_items pi WHERE pi.item_id=i.id)
          AND NOT EXISTS (SELECT 1 FROM return_items ri WHERE ri.item_id=i.id)
        ORDER BY i.name
        """,
        (business_id,),
    ).fetchall()
    return [dict(row) for row in rows if _clean_text(row["name"]) in parties]


@app.post("/api/import/cleanup-party-items")
def cleanup_party_items(
    execute: bool = Query(False),
    user: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner access required")
    with db() as conn:
        candidates = _wrong_party_item_candidates(conn, user["business_id"])
        if execute and candidates:
            ids = [row["id"] for row in candidates]
            placeholders = ",".join("?" for _ in ids)
            conn.execute(f"DELETE FROM stock_movements WHERE item_id IN ({placeholders})", ids)
            conn.execute(f"DELETE FROM items WHERE id IN ({placeholders})", ids)
    return {
        "execute": execute,
        "count": len(candidates),
        "items": [row["name"] for row in candidates[:50]],
    }


IMPORT_FIX_JS = r"""
(() => {
  'use strict';

  function detectImportType(headers, filename) {
    const file = String(filename || '').toLowerCase();
    if (file.includes('partyreport') || file.includes('party_report')) return 'parties';
    if (file.includes('purchasereport') || file.includes('purchase_report')) return 'purchases';
    if (file.includes('salereport') || file.includes('sale_report')) return 'sales';
    if (file.includes('item') || file.includes('stock')) return 'items';
    const set = new Set((headers || []).map(x => String(x).toLowerCase()));
    const has = (...names) => names.some(name => set.has(name));
    if (has('invoice_no','invoice_number','bill_no','transaction_no')) {
      if (has('supplier_name','purchase_price','purchase_rate')) return 'purchases';
      return 'sales';
    }
    if (has('party_name','mobile_number','opening_balance','current_balance','gstin')) return 'parties';
    if (has('item_name','product_name','sale_price','purchase_price','current_stock','opening_stock')) return 'items';
    return '';
  }

  function typeLabel(type) {
    return ({items:'Items / Stock',parties:'Parties',sales:'Sales',purchases:'Purchases'})[type] || type;
  }

  previewImport = async function safePreviewImport(doImport = false) {
    const file = $('#import-file')?.files?.[0];
    if (!file) return toast('File choose karein', true);
    const selected = $('#import-type').value;
    const previewData = new FormData();
    previewData.append('file', file);
    previewData.append('entity_type', selected);
    previewData.append('dry_run', 'true');
    try {
      const preview = await api('/api/import/vyapar', {method:'POST', body:previewData});
      const detected = detectImportType(preview.headers || [], file.name);
      if (detected && detected !== selected) {
        $('#import-result').classList.remove('hidden');
        $('#import-result').innerHTML = `<b style="color:#d33">Wrong data type selected</b><p>This file looks like <strong>${esc(typeLabel(detected))}</strong>, but you selected <strong>${esc(typeLabel(selected))}</strong>.</p><p>Data type change karke dobara import karein.</p>`;
        return toast(`Ye ${typeLabel(detected)} file hai`, true);
      }
      if (!doImport) {
        $('#import-result').classList.remove('hidden');
        $('#import-result').innerHTML = `<b>Preview ready</b><div class="simple-row"><span>Total rows</span><strong>${preview.rows_total || 0}</strong></div><div class="simple-row"><span>Detected</span><strong>${esc(typeLabel(detected || selected))}</strong></div><pre>${esc(JSON.stringify((preview.preview || []).slice(0,8), null, 2))}</pre>`;
        return;
      }
      const importData = new FormData();
      importData.append('file', file);
      importData.append('entity_type', selected);
      importData.append('dry_run', 'false');
      const result = await api('/api/import/vyapar', {method:'POST', body:importData});
      $('#import-result').classList.remove('hidden');
      const firstError = result.errors?.[0]?.error || '';
      $('#import-result').innerHTML = `<b>Import completed</b><div class="simple-row"><span>Total rows</span><strong>${result.rows_total || 0}</strong></div><div class="simple-row"><span>Imported</span><strong>${result.rows_imported || 0}</strong></div><div class="simple-row"><span>Skipped</span><strong>${result.rows_skipped || 0}</strong></div>${firstError ? `<p class="negative">${esc(firstError)}</p>` : ''}`;
      await refreshAll();
      await loadImportHistory();
    } catch (error) {
      toast(error.message, true);
    }
  };

  loadImportHistory = async function loadSafeImportHistory() {
    const [rows, cleanup] = await Promise.all([
      api('/api/import/batches'),
      api('/api/import/cleanup-party-items', {method:'POST'}),
    ]);
    const warning = cleanup.count ? `<div class="info-banner" style="margin-bottom:12px"><b>${cleanup.count} party names galti se Items mein mile</b><p>Sirf zero-stock, zero-rate aur unused imported records remove honge.</p><button id="cleanup-wrong-party-items" class="btn primary" type="button">Remove Wrong Party Items</button></div>` : '';
    $('#import-history').innerHTML = warning + (rows.map(row => {
      const firstError = row.errors?.[0]?.error || '';
      return `<div class="simple-row"><div><b>${esc(row.entity_type)} · ${esc(row.filename)}</b><small>${niceDate(row.created_at?.slice(0,10))}${firstError ? ` · ${esc(firstError)}` : ''}</small></div><strong>${row.rows_imported}/${row.rows_total}</strong></div>`;
    }).join('') || emptyText('Abhi koi import nahi hua.'));
  };

  document.addEventListener('click', async event => {
    const button = event.target.closest('#cleanup-wrong-party-items');
    if (!button) return;
    if (!confirm('Galti se bane party-name items remove karein? Correct items aur used bill items delete nahi honge.')) return;
    button.disabled = true;
    try {
      const result = await api('/api/import/cleanup-party-items?execute=true', {method:'POST'});
      toast(`${result.count} wrong items removed`);
      await refreshMasterData();
      renderItems();
      await loadImportHistory();
    } catch (error) {
      toast(error.message, true);
      button.disabled = false;
    }
  });
})();
"""


@app.get("/import-fix.js", include_in_schema=False)
def import_fix_javascript() -> Response:
    return Response(
        IMPORT_FIX_JS,
        media_type="application/javascript",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


@app.middleware("http")
async def inject_import_fix_assets(request, call_next):
    if request.method == "GET" and request.url.path == "/":
        html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
        html = html.replace("</head>", '<link rel="stylesheet" href="/settings-v2.css?v=046" /></head>')
        html = html.replace("</body>", '<script src="/settings-v2.js?v=046"></script><script src="/import-fix.js?v=046"></script></body>')
        return HTMLResponse(html, headers={"Cache-Control": "no-store"})
    return await call_next(request)


extension_paths = {"/api/import/cleanup-party-items", "/import-fix.js"}
extension_routes = [route for route in app.router.routes if getattr(route, "path", None) in extension_paths]
for route in extension_routes:
    app.router.routes.remove(route)
fallback_index = next(
    (index for index, route in enumerate(app.router.routes) if getattr(route, "path", None) == "/{path:path}"),
    len(app.router.routes),
)
app.router.routes[fallback_index:fallback_index] = extension_routes
