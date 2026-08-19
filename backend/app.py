from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Literal

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
try:
    import xlrd
except Exception:
    xlrd = None
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent.parent
STATIC_DIR = BASE_DIR / "static"
DB_PATH = Path(os.getenv("KIRANA_DB_PATH", str(BASE_DIR / "kirana.db")))

app = FastAPI(title="Kirana Software API", version="0.4.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def today_iso() -> str:
    return date.today().isoformat()


def money(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(",", "").replace("₹", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return default
    try:
        return round(float(text), 2)
    except ValueError:
        return default


def number(value: Any, default: float = 0.0) -> float:
    return money(value, default)


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_date(value: Any) -> str:
    if value is None or value == "":
        return today_iso()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return today_iso()


def normalize_size_label(number_text: str, unit_text: str = "") -> str:
    try:
        numeric = float(number_text)
        number_value = str(int(numeric)) if numeric.is_integer() else str(numeric).rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        number_value = str(number_text or "").strip()
    unit = str(unit_text or "").strip().lower().replace(".", "")
    unit_aliases = {
        "g": "gm", "gm": "gm", "gms": "gm", "gram": "gm", "grams": "gm",
        "kg": "kg", "kgs": "kg", "kilogram": "kg", "kilograms": "kg",
        "ml": "ml", "millilitre": "ml", "millilitres": "ml", "milliliter": "ml", "milliliters": "ml",
        "l": "ltr", "lt": "ltr", "ltr": "ltr", "litre": "ltr", "litres": "ltr", "liter": "ltr", "liters": "ltr",
        "pc": "pcs", "pcs": "pcs", "piece": "pcs", "pieces": "pcs",
        "pkt": "packet", "pkts": "packet", "packet": "packet", "packets": "packet",
    }
    normalized_unit = unit_aliases.get(unit, unit)
    return f"{number_value} {normalized_unit}".strip()


def split_item_name_size(name: Any, current_size: Any = "", item_unit: Any = "") -> tuple[str, str]:
    """Move an obvious pack size from an item name into the dedicated size field.

    Examples:
      Barik Souff 500 (बारिक सौंफ) -> Barik Souff (बारिक सौंफ), 500
      Sugar 1kg -> Sugar, 1 kg
    Plain trailing numbers below 10 are left untouched unless a unit is present,
    which avoids changing product names such as A2 or B12 accidentally.
    """
    raw_name = re.sub(r"\s+", " ", str(name or "")).strip()
    raw_size = re.sub(r"\s+", " ", str(current_size or "")).strip()
    if not raw_name:
        return raw_name, raw_size

    translation = ""
    core = raw_name
    translation_match = re.search(r"\s*(\([^()]*\))\s*$", raw_name)
    if translation_match:
        translation = translation_match.group(1)
        core = raw_name[:translation_match.start()].strip()

    unit_pattern = r"kg|kgs|kilogram|kilograms|g|gm|gms|gram|grams|ml|millilitre|millilitres|milliliter|milliliters|l|lt|ltr|litre|litres|liter|liters|pc|pcs|piece|pieces|pkt|pkts|packet|packets"
    match = re.match(rf"^(?P<base>.+?)\s+(?P<number>\d+(?:\.\d+)?)\s*(?P<unit>{unit_pattern})?$", core, flags=re.IGNORECASE)
    if not match:
        return raw_name, raw_size

    number_text = match.group("number")
    embedded_unit = match.group("unit") or ""
    try:
        numeric_value = float(number_text)
    except ValueError:
        return raw_name, raw_size

    # A unit makes the pack explicit. Without a unit, only common pack-like
    # numbers (10 and above) are moved to avoid corrupting model names.
    if not embedded_unit and numeric_value < 10:
        return raw_name, raw_size

    base = match.group("base").strip(" -_/,")
    if not base:
        return raw_name, raw_size
    clean_name = f"{base} {translation}".strip()
    detected_size = normalize_size_label(number_text, embedded_unit)
    return clean_name, detected_size or raw_size


@contextmanager
def db() -> Iterable[sqlite3.Connection]:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS businesses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                owner_name TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                gstin TEXT DEFAULT '',
                address TEXT DEFAULT '',
                invoice_prefix TEXT DEFAULT 'KS',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'owner',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                expires_at TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                sku TEXT DEFAULT '',
                barcode TEXT DEFAULT '',
                category TEXT DEFAULT '',
                unit TEXT DEFAULT 'pcs',
                size TEXT DEFAULT '',
                hsn TEXT DEFAULT '',
                gst_rate REAL NOT NULL DEFAULT 0,
                purchase_price REAL NOT NULL DEFAULT 0,
                sale_price REAL NOT NULL DEFAULT 0,
                mrp REAL NOT NULL DEFAULT 0,
                stock REAL NOT NULL DEFAULT 0,
                min_stock REAL NOT NULL DEFAULT 0,
                archived_at TEXT DEFAULT '',
                archived_reason TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(business_id, sku)
            );
            CREATE INDEX IF NOT EXISTS idx_items_business_name ON items(business_id, name);
            CREATE INDEX IF NOT EXISTS idx_items_barcode ON items(business_id, barcode);
            CREATE TABLE IF NOT EXISTS parties (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('customer','supplier','both')),
                phone TEXT DEFAULT '',
                gstin TEXT DEFAULT '',
                address TEXT DEFAULT '',
                opening_balance REAL NOT NULL DEFAULT 0,
                balance REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_parties_business_name ON parties(business_id, name);
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                invoice_no TEXT NOT NULL,
                party_id INTEGER REFERENCES parties(id) ON DELETE SET NULL,
                party_name TEXT DEFAULT 'Cash Customer',
                invoice_date TEXT NOT NULL,
                subtotal REAL NOT NULL,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL,
                paid REAL NOT NULL DEFAULT 0,
                due REAL NOT NULL DEFAULT 0,
                payment_mode TEXT DEFAULT 'cash',
                notes TEXT DEFAULT '',
                import_batch_id INTEGER,
                created_at TEXT NOT NULL,
                UNIQUE(business_id, invoice_no)
            );
            CREATE INDEX IF NOT EXISTS idx_sales_business_date ON sales(business_id, invoice_date DESC);
            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
                item_id INTEGER REFERENCES items(id) ON DELETE SET NULL,
                item_name TEXT NOT NULL,
                size TEXT DEFAULT '',
                qty REAL NOT NULL,
                rate REAL NOT NULL,
                gst_rate REAL NOT NULL DEFAULT 0,
                line_subtotal REAL NOT NULL,
                line_tax REAL NOT NULL,
                line_total REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS purchases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                invoice_no TEXT NOT NULL,
                party_id INTEGER REFERENCES parties(id) ON DELETE SET NULL,
                party_name TEXT DEFAULT 'Cash Supplier',
                invoice_date TEXT NOT NULL,
                subtotal REAL NOT NULL,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL,
                paid REAL NOT NULL DEFAULT 0,
                due REAL NOT NULL DEFAULT 0,
                payment_mode TEXT DEFAULT 'cash',
                notes TEXT DEFAULT '',
                import_batch_id INTEGER,
                created_at TEXT NOT NULL,
                UNIQUE(business_id, invoice_no)
            );
            CREATE INDEX IF NOT EXISTS idx_purchases_business_date ON purchases(business_id, invoice_date DESC);
            CREATE TABLE IF NOT EXISTS purchase_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                purchase_id INTEGER NOT NULL REFERENCES purchases(id) ON DELETE CASCADE,
                item_id INTEGER REFERENCES items(id) ON DELETE SET NULL,
                item_name TEXT NOT NULL,
                size TEXT DEFAULT '',
                qty REAL NOT NULL,
                rate REAL NOT NULL,
                gst_rate REAL NOT NULL DEFAULT 0,
                line_subtotal REAL NOT NULL,
                line_tax REAL NOT NULL,
                line_total REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                kind TEXT NOT NULL CHECK(kind IN ('sale_return','purchase_return')),
                return_no TEXT NOT NULL,
                party_id INTEGER REFERENCES parties(id) ON DELETE SET NULL,
                party_name TEXT DEFAULT '',
                return_date TEXT NOT NULL,
                reference_no TEXT DEFAULT '',
                subtotal REAL NOT NULL,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL,
                paid REAL NOT NULL DEFAULT 0,
                due REAL NOT NULL DEFAULT 0,
                payment_mode TEXT DEFAULT 'cash',
                notes TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                UNIQUE(business_id,kind,return_no)
            );
            CREATE INDEX IF NOT EXISTS idx_returns_business_date ON returns(business_id,return_date DESC);
            CREATE TABLE IF NOT EXISTS return_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                return_id INTEGER NOT NULL REFERENCES returns(id) ON DELETE CASCADE,
                item_id INTEGER REFERENCES items(id) ON DELETE SET NULL,
                item_name TEXT NOT NULL,
                size TEXT DEFAULT '',
                qty REAL NOT NULL,
                rate REAL NOT NULL,
                gst_rate REAL NOT NULL DEFAULT 0,
                line_subtotal REAL NOT NULL,
                line_tax REAL NOT NULL,
                line_total REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                item_id INTEGER NOT NULL REFERENCES items(id) ON DELETE CASCADE,
                movement_date TEXT NOT NULL,
                kind TEXT NOT NULL,
                qty REAL NOT NULL,
                reference_type TEXT DEFAULT '',
                reference_id INTEGER,
                note TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_stock_movements_item ON stock_movements(item_id, movement_date DESC);
            CREATE TABLE IF NOT EXISTS ledger_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                party_id INTEGER NOT NULL REFERENCES parties(id) ON DELETE CASCADE,
                entry_date TEXT NOT NULL,
                entry_type TEXT NOT NULL,
                reference_type TEXT DEFAULT '',
                reference_id INTEGER,
                debit REAL NOT NULL DEFAULT 0,
                credit REAL NOT NULL DEFAULT 0,
                note TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_ledger_party_date ON ledger_entries(party_id, entry_date DESC);
            CREATE TABLE IF NOT EXISTS import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                entity_type TEXT NOT NULL,
                filename TEXT NOT NULL,
                status TEXT NOT NULL,
                rows_total INTEGER NOT NULL DEFAULT 0,
                rows_imported INTEGER NOT NULL DEFAULT 0,
                rows_skipped INTEGER NOT NULL DEFAULT 0,
                errors_json TEXT DEFAULT '[]',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                account_type TEXT NOT NULL DEFAULT 'bank',
                balance REAL NOT NULL DEFAULT 0,
                is_default INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_accounts_business ON accounts(business_id, account_type);
            CREATE TABLE IF NOT EXISTS business_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                entry_type TEXT NOT NULL,
                entry_date TEXT NOT NULL,
                title TEXT DEFAULT '',
                party_id INTEGER REFERENCES parties(id) ON DELETE SET NULL,
                party_name TEXT DEFAULT '',
                account_id INTEGER REFERENCES accounts(id) ON DELETE SET NULL,
                to_account_id INTEGER REFERENCES accounts(id) ON DELETE SET NULL,
                amount REAL NOT NULL DEFAULT 0,
                status TEXT DEFAULT 'completed',
                mode TEXT DEFAULT 'cash',
                note TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_entries_business_date ON business_entries(business_id, entry_date DESC);
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_id INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
                kind TEXT NOT NULL,
                doc_no TEXT NOT NULL,
                doc_date TEXT NOT NULL,
                party_id INTEGER REFERENCES parties(id) ON DELETE SET NULL,
                party_name TEXT DEFAULT '',
                amount REAL NOT NULL DEFAULT 0,
                status TEXT DEFAULT 'open',
                note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                UNIQUE(business_id, kind, doc_no)
            );
            CREATE INDEX IF NOT EXISTS idx_documents_business_date ON documents(business_id, doc_date DESC);
            """
        )
        # Safe schema upgrades for existing shop databases.
        def ensure_column(table: str, column: str, definition: str) -> None:
            columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
            if column not in columns:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

        ensure_column("items", "size", "TEXT DEFAULT ''")
        ensure_column("items", "archived_at", "TEXT DEFAULT ''")
        ensure_column("items", "archived_reason", "TEXT DEFAULT ''")
        ensure_column("sale_items", "size", "TEXT DEFAULT ''")
        ensure_column("purchase_items", "size", "TEXT DEFAULT ''")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_items_business_archived "
            "ON items(business_id, archived_at)"
        )

        # One-time-safe cleanup for imported item names such as
        # “Barik Souff 500 (बारिक सौंफ)”. The pack is stored separately so
        # item lists and printed bills get a proper Size column.
        for row in conn.execute("SELECT id,name,size,unit FROM items").fetchall():
            clean_name, clean_size = split_item_name_size(row["name"], row["size"], row["unit"])
            if clean_name != row["name"] or clean_size != (row["size"] or ""):
                conn.execute(
                    "UPDATE items SET name=?,size=?,updated_at=? WHERE id=?",
                    (clean_name, clean_size, now_iso(), row["id"]),
                )
        for table in ("sale_items", "purchase_items", "return_items"):
            for row in conn.execute(f"SELECT id,item_name,size FROM {table}").fetchall():
                clean_name, clean_size = split_item_name_size(row["item_name"], row["size"], "")
                if clean_name != row["item_name"] or clean_size != (row["size"] or ""):
                    conn.execute(
                        f"UPDATE {table} SET item_name=?,size=? WHERE id=?",
                        (clean_name, clean_size, row["id"]),
                    )


@app.on_event("startup")
def startup() -> None:
    init_db()


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 120_000).hex()
    return f"{salt}${digest}"


def verify_password(password: str, encoded: str) -> bool:
    try:
        salt, expected = encoded.split("$", 1)
    except ValueError:
        return False
    actual = hash_password(password, salt).split("$", 1)[1]
    return hmac.compare_digest(actual, expected)


def rowdict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return dict(row) if row else None


class SetupIn(BaseModel):
    business_name: str = Field(min_length=2, max_length=120)
    owner_name: str = ""
    phone: str = ""
    gstin: str = ""
    address: str = ""
    username: str = Field(default="admin", min_length=3, max_length=40)
    password: str = Field(min_length=4, max_length=128)


class LoginIn(BaseModel):
    username: str
    password: str


class ItemIn(BaseModel):
    name: str = Field(min_length=1, max_length=160)
    sku: str = ""
    barcode: str = ""
    category: str = ""
    unit: str = "pcs"
    size: str = ""
    hsn: str = ""
    gst_rate: float = 0
    purchase_price: float = 0
    sale_price: float = 0
    mrp: float = 0
    stock: float = 0
    min_stock: float = 0


class PartyIn(BaseModel):
    name: str = Field(min_length=1, max_length=160)
    type: Literal["customer", "supplier", "both"] = "customer"
    phone: str = ""
    gstin: str = ""
    address: str = ""
    opening_balance: float = 0


class TxLineIn(BaseModel):
    item_id: int | None = None
    item_name: str = ""
    size: str = ""
    qty: float = Field(gt=0)
    rate: float = Field(ge=0)
    gst_rate: float = Field(default=0, ge=0)


class TransactionIn(BaseModel):
    invoice_no: str = ""
    party_id: int | None = None
    invoice_date: str = Field(default_factory=today_iso)
    discount: float = Field(default=0, ge=0)
    paid: float = Field(default=0, ge=0)
    payment_mode: str = "cash"
    notes: str = ""
    items: list[TxLineIn] = Field(min_items=1)


class ReturnIn(BaseModel):
    kind: Literal["sale_return", "purchase_return"]
    return_no: str = ""
    reference_no: str = ""
    party_id: int | None = None
    return_date: str = Field(default_factory=today_iso)
    discount: float = Field(default=0, ge=0)
    paid: float = Field(default=0, ge=0)
    payment_mode: str = "cash"
    notes: str = ""
    items: list[TxLineIn] = Field(min_items=1)


class PaymentIn(BaseModel):
    party_id: int
    amount: float = Field(gt=0)
    payment_type: Literal["received", "paid"]
    payment_date: str = Field(default_factory=today_iso)
    mode: str = "cash"
    note: str = ""


class AccountIn(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    account_type: Literal["cash", "bank", "loan", "asset"] = "bank"
    opening_balance: float = 0


class EntryIn(BaseModel):
    entry_type: Literal["expense", "cash_in", "cash_out", "transfer", "cheque_received", "cheque_paid", "loan_in", "loan_out", "asset_purchase", "asset_sale"]
    entry_date: str = Field(default_factory=today_iso)
    title: str = ""
    party_id: int | None = None
    account_id: int | None = None
    to_account_id: int | None = None
    amount: float = Field(gt=0)
    status: str = "completed"
    mode: str = "cash"
    note: str = ""


class DocumentIn(BaseModel):
    kind: Literal["sale_return", "purchase_return", "delivery_challan", "estimate", "proforma", "sale_order", "purchase_order", "sale_asset", "purchase_asset"]
    doc_no: str = ""
    doc_date: str = Field(default_factory=today_iso)
    party_id: int | None = None
    amount: float = Field(ge=0)
    status: str = "open"
    note: str = ""


class BusinessUpdateIn(BaseModel):
    name: str
    owner_name: str = ""
    phone: str = ""
    gstin: str = ""
    address: str = ""
    invoice_prefix: str = "KS"


def current_user(authorization: str | None = Header(default=None)) -> dict[str, Any]:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Login required")
    token = authorization.split(" ", 1)[1].strip()
    with db() as conn:
        row = conn.execute(
            """
            SELECT u.id AS user_id, u.business_id, u.username, u.role, b.name AS business_name
            FROM sessions s
            JOIN users u ON u.id=s.user_id
            JOIN businesses b ON b.id=u.business_id
            WHERE s.token=? AND s.expires_at>?
            """,
            (token, now_iso()),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Session expired")
    return dict(row)


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {"ok": True, "name": "Kirana Software", "version": app.version, "time": now_iso()}


@app.get("/api/setup/status")
def setup_status() -> dict[str, Any]:
    with db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0]
    return {"setup_complete": count > 0}


@app.post("/api/setup")
def setup(payload: SetupIn) -> dict[str, Any]:
    with db() as conn:
        if conn.execute("SELECT COUNT(*) FROM businesses").fetchone()[0] > 0:
            raise HTTPException(status_code=409, detail="Software is already set up")
        cur = conn.execute(
            "INSERT INTO businesses(name,owner_name,phone,gstin,address,invoice_prefix,created_at) VALUES(?,?,?,?,?,?,?)",
            (payload.business_name, payload.owner_name, payload.phone, payload.gstin, payload.address, "KS", now_iso()),
        )
        business_id = int(cur.lastrowid)
        conn.execute(
            "INSERT INTO accounts(business_id,name,account_type,balance,is_default,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
            (business_id, "Cash In Hand", "cash", 0, 1, now_iso(), now_iso()),
        )
        cur = conn.execute(
            "INSERT INTO users(business_id,username,password_hash,role,created_at) VALUES(?,?,?,?,?)",
            (business_id, payload.username.strip().lower(), hash_password(payload.password), "owner", now_iso()),
        )
        user_id = int(cur.lastrowid)
        token = secrets.token_urlsafe(40)
        expires_at = (datetime.now() + timedelta(days=30)).replace(microsecond=0).isoformat()
        conn.execute(
            "INSERT INTO sessions(token,user_id,expires_at,created_at) VALUES(?,?,?,?)",
            (token, user_id, expires_at, now_iso()),
        )
    return {"token": token, "business_id": business_id, "business_name": payload.business_name}


@app.post("/api/login")
def login(payload: LoginIn) -> dict[str, Any]:
    with db() as conn:
        row = conn.execute(
            "SELECT u.*, b.name AS business_name FROM users u JOIN businesses b ON b.id=u.business_id WHERE u.username=?",
            (payload.username.strip().lower(),),
        ).fetchone()
        if not row or not verify_password(payload.password, row["password_hash"]):
            raise HTTPException(status_code=401, detail="Wrong username or password")
        token = secrets.token_urlsafe(40)
        expires_at = (datetime.now() + timedelta(days=30)).replace(microsecond=0).isoformat()
        conn.execute("DELETE FROM sessions WHERE expires_at<=?", (now_iso(),))
        conn.execute(
            "INSERT INTO sessions(token,user_id,expires_at,created_at) VALUES(?,?,?,?)",
            (token, row["id"], expires_at, now_iso()),
        )
    return {"token": token, "business_id": row["business_id"], "business_name": row["business_name"]}


@app.post("/api/logout")
def logout(user: dict[str, Any] = Depends(current_user), authorization: str | None = Header(default=None)) -> dict[str, bool]:
    token = authorization.split(" ", 1)[1].strip() if authorization else ""
    with db() as conn:
        conn.execute("DELETE FROM sessions WHERE token=?", (token,))
    return {"ok": True}


@app.get("/api/me")
def me(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        business = rowdict(conn.execute("SELECT * FROM businesses WHERE id=?", (user["business_id"],)).fetchone())
    return {"user": user, "business": business}


@app.put("/api/business")
def update_business(payload: BusinessUpdateIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        conn.execute(
            "UPDATE businesses SET name=?,owner_name=?,phone=?,gstin=?,address=?,invoice_prefix=? WHERE id=?",
            (payload.name, payload.owner_name, payload.phone, payload.gstin, payload.address, payload.invoice_prefix.upper()[:8], user["business_id"]),
        )
        business = rowdict(conn.execute("SELECT * FROM businesses WHERE id=?", (user["business_id"],)).fetchone())
    return business or {}


def next_invoice(conn: sqlite3.Connection, business_id: int, table: str, prefix_suffix: str) -> str:
    business = conn.execute("SELECT invoice_prefix FROM businesses WHERE id=?", (business_id,)).fetchone()
    prefix = (business["invoice_prefix"] or "KS").upper()
    fy = f"{date.today().year % 100:02d}{(date.today().year + 1) % 100:02d}"
    base = f"{prefix}-{prefix_suffix}-{fy}-"
    row = conn.execute(
        f"SELECT invoice_no FROM {table} WHERE business_id=? AND invoice_no LIKE ? ORDER BY id DESC LIMIT 1",
        (business_id, f"{base}%"),
    ).fetchone()
    seq = 1
    if row:
        try:
            seq = int(str(row["invoice_no"]).split("-")[-1]) + 1
        except ValueError:
            pass
    return f"{base}{seq:05d}"


@app.get("/api/dashboard")
def dashboard(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid = user["business_id"]
    today = today_iso()
    month_start = date.today().replace(day=1).isoformat()
    year = date.today().year
    month = date.today().month
    trend_start = (date.today().replace(day=1) - timedelta(days=155)).replace(day=1)
    with db() as conn:
        # Keep old databases compatible by ensuring a default cash account exists.
        if conn.execute("SELECT COUNT(*) FROM accounts WHERE business_id=?", (bid,)).fetchone()[0] == 0:
            conn.execute("INSERT INTO accounts(business_id,name,account_type,balance,is_default,created_at,updated_at) VALUES(?,?,?,?,?,?,?)", (bid, "Cash In Hand", "cash", 0, 1, now_iso(), now_iso()))
        sales_today = conn.execute("SELECT COALESCE(SUM(total),0) FROM sales WHERE business_id=? AND invoice_date=?", (bid, today)).fetchone()[0]
        purchases_today = conn.execute("SELECT COALESCE(SUM(total),0) FROM purchases WHERE business_id=? AND invoice_date=?", (bid, today)).fetchone()[0]
        sales_month = conn.execute("SELECT COALESCE(SUM(total),0) FROM sales WHERE business_id=? AND invoice_date>=?", (bid, month_start)).fetchone()[0]
        purchases_month = conn.execute("SELECT COALESCE(SUM(total),0) FROM purchases WHERE business_id=? AND invoice_date>=?", (bid, month_start)).fetchone()[0]
        expenses_month = conn.execute("SELECT COALESCE(SUM(amount),0) FROM business_entries WHERE business_id=? AND entry_type='expense' AND entry_date>=?", (bid, month_start)).fetchone()[0]
        receivable = conn.execute("SELECT COALESCE(SUM(balance),0) FROM parties WHERE business_id=? AND type IN ('customer','both') AND balance>0", (bid,)).fetchone()[0]
        payable = conn.execute("SELECT COALESCE(SUM(balance),0) FROM parties WHERE business_id=? AND type IN ('supplier','both') AND balance>0", (bid,)).fetchone()[0]
        active_item = "COALESCE(archived_at,'')=''"
        low_stock = conn.execute(f"SELECT COUNT(*) FROM items WHERE business_id=? AND {active_item} AND stock<=min_stock", (bid,)).fetchone()[0]
        item_count = conn.execute(f"SELECT COUNT(*) FROM items WHERE business_id=? AND {active_item}", (bid,)).fetchone()[0]
        stock_value = conn.execute(f"SELECT COALESCE(SUM(stock*purchase_price),0) FROM items WHERE business_id=? AND {active_item}", (bid,)).fetchone()[0]
        accounts = [dict(r) for r in conn.execute("SELECT * FROM accounts WHERE business_id=? ORDER BY is_default DESC,name", (bid,)).fetchall()]
        bank_balance = round(sum(float(r["balance"]) for r in accounts if r["account_type"] == "bank"), 2)
        cash_balance = round(sum(float(r["balance"]) for r in accounts if r["account_type"] == "cash"), 2)
        recent_sales = [dict(r) for r in conn.execute("SELECT id,invoice_no,party_name,invoice_date,total,paid,due,'sale' AS kind FROM sales WHERE business_id=? ORDER BY id DESC LIMIT 8", (bid,)).fetchall()]
        recent_purchases = [dict(r) for r in conn.execute("SELECT id,invoice_no,party_name,invoice_date,total,paid,due,'purchase' AS kind FROM purchases WHERE business_id=? ORDER BY id DESC LIMIT 8", (bid,)).fetchall()]
        recent_entries = [dict(r) for r in conn.execute("SELECT id,title AS invoice_no,COALESCE(party_name,title) AS party_name,entry_date AS invoice_date,amount AS total,amount AS paid,0 AS due,entry_type AS kind,status FROM business_entries WHERE business_id=? ORDER BY id DESC LIMIT 10", (bid,)).fetchall()]
        activity = sorted(recent_sales + recent_purchases + recent_entries, key=lambda x: (x.get("invoice_date", ""), x.get("id", 0)), reverse=True)[:12]
        low_items = [dict(r) for r in conn.execute(f"SELECT id,name,stock,min_stock,unit FROM items WHERE business_id=? AND {active_item} AND stock<=min_stock ORDER BY stock ASC LIMIT 5", (bid,)).fetchall()]
        open_docs = conn.execute("SELECT COUNT(*),COALESCE(SUM(amount),0) FROM documents WHERE business_id=? AND status NOT IN ('closed','cancelled')", (bid,)).fetchone()
        trend_rows = conn.execute("SELECT substr(invoice_date,1,7) AS ym,COALESCE(SUM(total),0) AS amount FROM sales WHERE business_id=? AND invoice_date>=? GROUP BY ym ORDER BY ym", (bid, trend_start.isoformat())).fetchall()
    trend_map = {r["ym"]: round(r["amount"],2) for r in trend_rows}
    trend=[]
    cursor=trend_start
    for _ in range(6):
        ym=cursor.strftime("%Y-%m")
        trend.append({"month":cursor.strftime("%b"),"year":cursor.year,"amount":trend_map.get(ym,0)})
        cursor=(cursor.replace(day=28)+timedelta(days=4)).replace(day=1)
    return {
        "sales_today": round(sales_today, 2), "purchases_today": round(purchases_today, 2),
        "sales_month": round(sales_month, 2), "purchases_month": round(purchases_month, 2),
        "expenses_month": round(expenses_month, 2), "receivable": round(receivable, 2),
        "payable": round(payable, 2), "low_stock": low_stock, "item_count": item_count,
        "stock_value": round(stock_value,2), "bank_balance": bank_balance, "cash_balance": cash_balance,
        "recent_sales": recent_sales, "activity": activity, "accounts": accounts, "low_items": low_items,
        "open_documents": {"count": open_docs[0], "amount": round(open_docs[1],2)}, "sales_trend": trend,
    }


@app.get("/api/items")
def list_items(
    q: str = "",
    low_stock: bool = False,
    include_archived: bool = False,
    limit: int = Query(default=500, ge=1, le=2000),
    user: dict[str, Any] = Depends(current_user),
) -> list[dict[str, Any]]:
    bid = user["business_id"]
    sql = "SELECT * FROM items WHERE business_id=?"
    args: list[Any] = [bid]
    if not include_archived:
        sql += " AND COALESCE(archived_at,'')=''"
    if q:
        sql += " AND (name LIKE ? OR size LIKE ? OR sku LIKE ? OR barcode LIKE ? OR category LIKE ?)"
        like = f"%{q}%"
        args += [like, like, like, like, like]
    if low_stock:
        sql += " AND stock<=min_stock"
    sql += " ORDER BY name,size LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(r) for r in conn.execute(sql, args).fetchall()]


@app.post("/api/items")
def create_item(payload: ItemIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid = user["business_id"]
    sku = payload.sku.strip() or f"ITEM-{secrets.token_hex(3).upper()}"
    clean_name, clean_size = split_item_name_size(payload.name, payload.size, payload.unit)
    with db() as conn:
        try:
            cur = conn.execute(
                """
                INSERT INTO items(business_id,name,sku,barcode,category,unit,size,hsn,gst_rate,purchase_price,sale_price,mrp,stock,min_stock,created_at,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (bid, clean_name, sku, payload.barcode.strip(), payload.category.strip(), payload.unit.strip() or "pcs", clean_size, payload.hsn.strip(), payload.gst_rate, payload.purchase_price, payload.sale_price, payload.mrp, payload.stock, payload.min_stock, now_iso(), now_iso()),
            )
        except sqlite3.IntegrityError as exc:
            raise HTTPException(status_code=409, detail="SKU already exists") from exc
        item_id = int(cur.lastrowid)
        if payload.stock:
            conn.execute(
                "INSERT INTO stock_movements(business_id,item_id,movement_date,kind,qty,reference_type,note,created_at) VALUES(?,?,?,?,?,?,?,?)",
                (bid, item_id, today_iso(), "opening", payload.stock, "opening", "Opening stock", now_iso()),
            )
        return dict(conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone())


@app.put("/api/items/{item_id}")
def update_item(item_id: int, payload: ItemIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid = user["business_id"]
    with db() as conn:
        old = conn.execute("SELECT * FROM items WHERE id=? AND business_id=?", (item_id, bid)).fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Item not found")
        sku = payload.sku.strip() or old["sku"]
        clean_name, clean_size = split_item_name_size(payload.name, payload.size, payload.unit)
        try:
            conn.execute(
                """
                UPDATE items SET name=?,sku=?,barcode=?,category=?,unit=?,size=?,hsn=?,gst_rate=?,purchase_price=?,sale_price=?,mrp=?,stock=?,min_stock=?,updated_at=?
                WHERE id=? AND business_id=?
                """,
                (clean_name, sku, payload.barcode.strip(), payload.category.strip(), payload.unit.strip() or "pcs", clean_size, payload.hsn.strip(), payload.gst_rate, payload.purchase_price, payload.sale_price, payload.mrp, payload.stock, payload.min_stock, now_iso(), item_id, bid),
            )
        except sqlite3.IntegrityError as exc:
            raise HTTPException(status_code=409, detail="SKU already exists") from exc
        difference = round(payload.stock - old["stock"], 4)
        if difference:
            conn.execute(
                "INSERT INTO stock_movements(business_id,item_id,movement_date,kind,qty,reference_type,note,created_at) VALUES(?,?,?,?,?,?,?,?)",
                (bid, item_id, today_iso(), "adjustment", difference, "manual", "Manual stock adjustment", now_iso()),
            )
        return dict(conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone())


@app.delete("/api/items/{item_id}")
def delete_item(item_id: int, user: dict[str, Any] = Depends(current_user)) -> dict[str, bool]:
    with db() as conn:
        exists = conn.execute("SELECT id FROM items WHERE id=? AND business_id=?", (item_id, user["business_id"])).fetchone()
        if not exists:
            raise HTTPException(status_code=404, detail="Item not found")
        used = conn.execute(
            "SELECT 1 FROM sale_items WHERE item_id=? UNION SELECT 1 FROM purchase_items WHERE item_id=? UNION SELECT 1 FROM return_items WHERE item_id=? LIMIT 1",
            (item_id, item_id, item_id),
        ).fetchone()
        if used:
            raise HTTPException(status_code=409, detail="Item has transactions and cannot be deleted")
        conn.execute("DELETE FROM items WHERE id=?", (item_id,))
    return {"ok": True}


@app.get("/api/items/{item_id}/movements")
def item_movements(item_id: int, user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    with db() as conn:
        item = conn.execute("SELECT id FROM items WHERE id=? AND business_id=?", (item_id, user["business_id"])).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        return [dict(r) for r in conn.execute("SELECT * FROM stock_movements WHERE item_id=? ORDER BY id DESC LIMIT 200", (item_id,)).fetchall()]


@app.get("/api/parties")
def list_parties(
    q: str = "",
    party_type: str = "",
    user: dict[str, Any] = Depends(current_user),
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM parties WHERE business_id=?"
    args: list[Any] = [user["business_id"]]
    if q:
        like = f"%{q}%"
        sql += " AND (name LIKE ? OR phone LIKE ? OR gstin LIKE ?)"
        args += [like, like, like]
    if party_type in {"customer", "supplier", "both"}:
        sql += " AND type IN (?, 'both')" if party_type != "both" else " AND type='both'"
        args.append(party_type) if party_type != "both" else None
    sql += " ORDER BY name"
    with db() as conn:
        return [dict(r) for r in conn.execute(sql, args).fetchall()]


@app.post("/api/parties")
def create_party(payload: PartyIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO parties(business_id,name,type,phone,gstin,address,opening_balance,balance,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (user["business_id"], payload.name.strip(), payload.type, payload.phone.strip(), payload.gstin.strip(), payload.address.strip(), payload.opening_balance, payload.opening_balance, now_iso(), now_iso()),
        )
        party_id = int(cur.lastrowid)
        if payload.opening_balance:
            conn.execute(
                "INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (user["business_id"], party_id, today_iso(), "opening", "opening", payload.opening_balance, 0, "Opening balance", now_iso()),
            )
        return dict(conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone())


@app.put("/api/parties/{party_id}")
def update_party(party_id: int, payload: PartyIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        old = conn.execute("SELECT * FROM parties WHERE id=? AND business_id=?", (party_id, user["business_id"])).fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Party not found")
        balance = old["balance"] + (payload.opening_balance - old["opening_balance"])
        conn.execute(
            "UPDATE parties SET name=?,type=?,phone=?,gstin=?,address=?,opening_balance=?,balance=?,updated_at=? WHERE id=?",
            (payload.name.strip(), payload.type, payload.phone.strip(), payload.gstin.strip(), payload.address.strip(), payload.opening_balance, balance, now_iso(), party_id),
        )
        return dict(conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone())


@app.get("/api/parties/{party_id}/ledger")
def party_ledger(party_id: int, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        party = conn.execute("SELECT * FROM parties WHERE id=? AND business_id=?", (party_id, user["business_id"])).fetchone()
        if not party:
            raise HTTPException(status_code=404, detail="Party not found")
        entries = [dict(r) for r in conn.execute("SELECT * FROM ledger_entries WHERE party_id=? ORDER BY entry_date DESC,id DESC LIMIT 500", (party_id,)).fetchall()]
    return {"party": dict(party), "entries": entries}


def account_for_mode(conn: sqlite3.Connection, bid: int, mode: str) -> int | None:
    wanted = "bank" if str(mode).lower() in {"bank", "card", "upi"} else "cash"
    row = conn.execute("SELECT id FROM accounts WHERE business_id=? AND account_type=? ORDER BY is_default DESC,id LIMIT 1", (bid, wanted)).fetchone()
    if not row:
        row = conn.execute("SELECT id FROM accounts WHERE business_id=? ORDER BY is_default DESC,id LIMIT 1", (bid,)).fetchone()
    return int(row["id"]) if row else None


def adjust_account(conn: sqlite3.Connection, bid: int, mode: str, amount: float) -> None:
    if not amount:
        return
    account_id = account_for_mode(conn, bid, mode)
    if account_id:
        conn.execute("UPDATE accounts SET balance=balance+?,updated_at=? WHERE id=? AND business_id=?", (amount, now_iso(), account_id, bid))


def calculate_lines(conn: sqlite3.Connection, bid: int, lines: list[TxLineIn], rate_field: str) -> tuple[list[dict[str, Any]], float, float]:
    prepared: list[dict[str, Any]] = []
    subtotal = 0.0
    tax = 0.0
    for line in lines:
        item = None
        if line.item_id:
            item = conn.execute("SELECT * FROM items WHERE id=? AND business_id=?", (line.item_id, bid)).fetchone()
            if not item:
                raise HTTPException(status_code=404, detail=f"Item {line.item_id} not found")
        item_name = line.item_name.strip() or (item["name"] if item else "Item")
        size = line.size.strip() or (str(item["size"] or "") if item else "")
        item_name, size = split_item_name_size(item_name, size, item["unit"] if item else "")
        rate = line.rate if line.rate is not None else (item[rate_field] if item else 0)
        gst = line.gst_rate if line.gst_rate is not None else (item["gst_rate"] if item else 0)
        line_subtotal = round(line.qty * rate, 2)
        line_tax = round(line_subtotal * gst / 100, 2)
        line_total = round(line_subtotal + line_tax, 2)
        subtotal += line_subtotal
        tax += line_tax
        prepared.append({
            "item_id": line.item_id,
            "item_name": item_name,
            "size": size,
            "qty": line.qty,
            "rate": rate,
            "gst_rate": gst,
            "line_subtotal": line_subtotal,
            "line_tax": line_tax,
            "line_total": line_total,
        })
    return prepared, round(subtotal, 2), round(tax, 2)


def get_party(conn: sqlite3.Connection, bid: int, party_id: int | None, expected: str) -> sqlite3.Row | None:
    if party_id is None:
        return None
    party = conn.execute("SELECT * FROM parties WHERE id=? AND business_id=?", (party_id, bid)).fetchone()
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    if party["type"] not in {expected, "both"}:
        raise HTTPException(status_code=400, detail=f"Selected party is not a {expected}")
    return party


def insert_sale(conn: sqlite3.Connection, bid: int, payload: TransactionIn, import_batch_id: int | None = None) -> dict[str, Any]:
    party = get_party(conn, bid, payload.party_id, "customer")
    lines, subtotal, tax = calculate_lines(conn, bid, payload.items, "sale_price")
    total = max(0.0, round(subtotal + tax - payload.discount, 2))
    paid = min(round(payload.paid, 2), total)
    due = round(total - paid, 2)
    invoice_no = payload.invoice_no.strip() or next_invoice(conn, bid, "sales", "S")
    try:
        cur = conn.execute(
            """
            INSERT INTO sales(business_id,invoice_no,party_id,party_name,invoice_date,subtotal,discount,tax,total,paid,due,payment_mode,notes,import_batch_id,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (bid, invoice_no, payload.party_id, party["name"] if party else "Cash Customer", normalize_date(payload.invoice_date), subtotal, payload.discount, tax, total, paid, due, payload.payment_mode, payload.notes, import_batch_id, now_iso()),
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail=f"Sale invoice {invoice_no} already exists") from exc
    sale_id = int(cur.lastrowid)
    for line in lines:
        conn.execute(
            "INSERT INTO sale_items(sale_id,item_id,item_name,size,qty,rate,gst_rate,line_subtotal,line_tax,line_total) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (sale_id, line["item_id"], line["item_name"], line["size"], line["qty"], line["rate"], line["gst_rate"], line["line_subtotal"], line["line_tax"], line["line_total"]),
        )
        if line["item_id"]:
            conn.execute("UPDATE items SET stock=stock-?,updated_at=? WHERE id=?", (line["qty"], now_iso(), line["item_id"]))
            conn.execute(
                "INSERT INTO stock_movements(business_id,item_id,movement_date,kind,qty,reference_type,reference_id,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (bid, line["item_id"], normalize_date(payload.invoice_date), "sale", -line["qty"], "sale", sale_id, invoice_no, now_iso()),
            )
    adjust_account(conn, bid, payload.payment_mode, paid)
    if party and due:
        conn.execute("UPDATE parties SET balance=balance+?,updated_at=? WHERE id=?", (due, now_iso(), party["id"]))
        conn.execute(
            "INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,reference_id,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (bid, party["id"], normalize_date(payload.invoice_date), "sale", "sale", sale_id, due, 0, invoice_no, now_iso()),
        )
    return transaction_detail(conn, "sale", sale_id, bid)


def insert_purchase(conn: sqlite3.Connection, bid: int, payload: TransactionIn, import_batch_id: int | None = None) -> dict[str, Any]:
    party = get_party(conn, bid, payload.party_id, "supplier")
    lines, subtotal, tax = calculate_lines(conn, bid, payload.items, "purchase_price")
    total = max(0.0, round(subtotal + tax - payload.discount, 2))
    paid = min(round(payload.paid, 2), total)
    due = round(total - paid, 2)
    invoice_no = payload.invoice_no.strip() or next_invoice(conn, bid, "purchases", "P")
    try:
        cur = conn.execute(
            """
            INSERT INTO purchases(business_id,invoice_no,party_id,party_name,invoice_date,subtotal,discount,tax,total,paid,due,payment_mode,notes,import_batch_id,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (bid, invoice_no, payload.party_id, party["name"] if party else "Cash Supplier", normalize_date(payload.invoice_date), subtotal, payload.discount, tax, total, paid, due, payload.payment_mode, payload.notes, import_batch_id, now_iso()),
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail=f"Purchase invoice {invoice_no} already exists") from exc
    purchase_id = int(cur.lastrowid)
    for line in lines:
        conn.execute(
            "INSERT INTO purchase_items(purchase_id,item_id,item_name,size,qty,rate,gst_rate,line_subtotal,line_tax,line_total) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (purchase_id, line["item_id"], line["item_name"], line["size"], line["qty"], line["rate"], line["gst_rate"], line["line_subtotal"], line["line_tax"], line["line_total"]),
        )
        if line["item_id"]:
            conn.execute("UPDATE items SET stock=stock+?,purchase_price=?,updated_at=? WHERE id=?", (line["qty"], line["rate"], now_iso(), line["item_id"]))
            conn.execute(
                "INSERT INTO stock_movements(business_id,item_id,movement_date,kind,qty,reference_type,reference_id,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (bid, line["item_id"], normalize_date(payload.invoice_date), "purchase", line["qty"], "purchase", purchase_id, invoice_no, now_iso()),
            )
    adjust_account(conn, bid, payload.payment_mode, -paid)
    if party and due:
        conn.execute("UPDATE parties SET balance=balance+?,updated_at=? WHERE id=?", (due, now_iso(), party["id"]))
        conn.execute(
            "INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,reference_id,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (bid, party["id"], normalize_date(payload.invoice_date), "purchase", "purchase", purchase_id, due, 0, invoice_no, now_iso()),
        )
    return transaction_detail(conn, "purchase", purchase_id, bid)


def transaction_detail(conn: sqlite3.Connection, kind: str, tx_id: int, bid: int) -> dict[str, Any]:
    table = "sales" if kind == "sale" else "purchases"
    line_table = "sale_items" if kind == "sale" else "purchase_items"
    fk = "sale_id" if kind == "sale" else "purchase_id"
    tx = conn.execute(f"SELECT * FROM {table} WHERE id=? AND business_id=?", (tx_id, bid)).fetchone()
    if not tx:
        raise HTTPException(status_code=404, detail=f"{kind.title()} not found")
    lines = [dict(r) for r in conn.execute(f"SELECT * FROM {line_table} WHERE {fk}=?", (tx_id,)).fetchall()]
    result = dict(tx)
    result["items"] = lines
    return result


@app.post("/api/sales")
def create_sale(payload: TransactionIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        return insert_sale(conn, user["business_id"], payload)


@app.get("/api/sales")
def list_sales(
    q: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = Query(default=200, ge=1, le=1000),
    user: dict[str, Any] = Depends(current_user),
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM sales WHERE business_id=?"
    args: list[Any] = [user["business_id"]]
    if q:
        like = f"%{q}%"
        sql += " AND (invoice_no LIKE ? OR party_name LIKE ?)"
        args += [like, like]
    if date_from:
        sql += " AND invoice_date>=?"
        args.append(normalize_date(date_from))
    if date_to:
        sql += " AND invoice_date<=?"
        args.append(normalize_date(date_to))
    sql += " ORDER BY invoice_date DESC,id DESC LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(r) for r in conn.execute(sql, args).fetchall()]


@app.get("/api/sales/{sale_id}")
def get_sale(sale_id: int, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        result = transaction_detail(conn, "sale", sale_id, user["business_id"])
        result["business"] = rowdict(conn.execute("SELECT * FROM businesses WHERE id=?", (user["business_id"],)).fetchone())
        if result.get("party_id"):
            result["party"] = rowdict(conn.execute("SELECT * FROM parties WHERE id=?", (result["party_id"],)).fetchone())
    return result


@app.post("/api/purchases")
def create_purchase(payload: TransactionIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        return insert_purchase(conn, user["business_id"], payload)


@app.get("/api/purchases")
def list_purchases(
    q: str = "",
    date_from: str = "",
    date_to: str = "",
    limit: int = Query(default=200, ge=1, le=1000),
    user: dict[str, Any] = Depends(current_user),
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM purchases WHERE business_id=?"
    args: list[Any] = [user["business_id"]]
    if q:
        like = f"%{q}%"
        sql += " AND (invoice_no LIKE ? OR party_name LIKE ?)"
        args += [like, like]
    if date_from:
        sql += " AND invoice_date>=?"
        args.append(normalize_date(date_from))
    if date_to:
        sql += " AND invoice_date<=?"
        args.append(normalize_date(date_to))
    sql += " ORDER BY invoice_date DESC,id DESC LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(r) for r in conn.execute(sql, args).fetchall()]


@app.get("/api/purchases/{purchase_id}")
def get_purchase(purchase_id: int, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        result = transaction_detail(conn, "purchase", purchase_id, user["business_id"])
        result["business"] = rowdict(conn.execute("SELECT * FROM businesses WHERE id=?", (user["business_id"],)).fetchone())
        if result.get("party_id"):
            result["party"] = rowdict(conn.execute("SELECT * FROM parties WHERE id=?", (result["party_id"],)).fetchone())
        return result


def next_return_no(conn: sqlite3.Connection, bid: int, kind: str) -> str:
    prefix = "SR" if kind == "sale_return" else "PR"
    row = conn.execute(
        "SELECT id FROM returns WHERE business_id=? AND kind=? ORDER BY id DESC LIMIT 1",
        (bid, kind),
    ).fetchone()
    seq = (int(row["id"]) + 1) if row else 1
    return f"{prefix}-{date.today().year % 100:02d}-{seq:05d}"


def return_detail(conn: sqlite3.Connection, return_id: int, bid: int) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM returns WHERE id=? AND business_id=?", (return_id, bid)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Return not found")
    result = dict(row)
    result["items"] = [dict(x) for x in conn.execute("SELECT * FROM return_items WHERE return_id=? ORDER BY id", (return_id,)).fetchall()]
    return result


def insert_return(conn: sqlite3.Connection, bid: int, payload: ReturnIn) -> dict[str, Any]:
    is_sale = payload.kind == "sale_return"
    party = get_party(conn, bid, payload.party_id, "customer" if is_sale else "supplier")
    lines, subtotal, tax = calculate_lines(conn, bid, payload.items, "sale_price" if is_sale else "purchase_price")
    total = max(0.0, round(subtotal + tax - payload.discount, 2))
    paid = min(round(payload.paid, 2), total)
    due = round(total - paid, 2)
    return_no = payload.return_no.strip() or next_return_no(conn, bid, payload.kind)
    try:
        cur = conn.execute(
            """
            INSERT INTO returns(business_id,kind,return_no,party_id,party_name,return_date,reference_no,subtotal,discount,tax,total,paid,due,payment_mode,notes,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (bid, payload.kind, return_no, payload.party_id, party["name"] if party else ("Cash Customer" if is_sale else "Cash Supplier"), normalize_date(payload.return_date), payload.reference_no.strip(), subtotal, payload.discount, tax, total, paid, due, payload.payment_mode, payload.notes, now_iso()),
        )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail=f"Return number {return_no} already exists") from exc
    return_id = int(cur.lastrowid)
    stock_sign = 1 if is_sale else -1
    for line in lines:
        conn.execute(
            "INSERT INTO return_items(return_id,item_id,item_name,size,qty,rate,gst_rate,line_subtotal,line_tax,line_total) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (return_id, line["item_id"], line["item_name"], line["size"], line["qty"], line["rate"], line["gst_rate"], line["line_subtotal"], line["line_tax"], line["line_total"]),
        )
        if line["item_id"]:
            delta = stock_sign * line["qty"]
            conn.execute("UPDATE items SET stock=stock+?,updated_at=? WHERE id=?", (delta, now_iso(), line["item_id"]))
            conn.execute(
                "INSERT INTO stock_movements(business_id,item_id,movement_date,kind,qty,reference_type,reference_id,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (bid, line["item_id"], normalize_date(payload.return_date), payload.kind, delta, payload.kind, return_id, return_no, now_iso()),
            )
    # paid means an actual refund settlement. Remaining value adjusts party outstanding.
    adjust_account(conn, bid, payload.payment_mode, -paid if is_sale else paid)
    if party and due:
        new_balance = max(0.0, round(float(party["balance"]) - due, 2))
        conn.execute("UPDATE parties SET balance=?,updated_at=? WHERE id=?", (new_balance, now_iso(), party["id"]))
        conn.execute(
            "INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,reference_id,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (bid, party["id"], normalize_date(payload.return_date), payload.kind, payload.kind, return_id, 0, due, f"{return_no} {payload.reference_no}".strip(), now_iso()),
        )
    return return_detail(conn, return_id, bid)


@app.post("/api/returns")
def create_return(payload: ReturnIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        return insert_return(conn, user["business_id"], payload)


@app.get("/api/returns")
def list_returns(
    kind: str = "",
    q: str = "",
    limit: int = Query(default=200, ge=1, le=1000),
    user: dict[str, Any] = Depends(current_user),
) -> list[dict[str, Any]]:
    sql = "SELECT * FROM returns WHERE business_id=?"
    args: list[Any] = [user["business_id"]]
    if kind in {"sale_return", "purchase_return"}:
        sql += " AND kind=?"
        args.append(kind)
    if q:
        like = f"%{q}%"
        sql += " AND (return_no LIKE ? OR party_name LIKE ? OR reference_no LIKE ?)"
        args += [like, like, like]
    sql += " ORDER BY return_date DESC,id DESC LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(row) for row in conn.execute(sql, args).fetchall()]


@app.get("/api/returns/{return_id}")
def get_return(return_id: int, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        result = return_detail(conn, return_id, user["business_id"])
        result["business"] = rowdict(conn.execute("SELECT * FROM businesses WHERE id=?", (user["business_id"],)).fetchone())
        if result.get("party_id"):
            result["party"] = rowdict(conn.execute("SELECT * FROM parties WHERE id=?", (result["party_id"],)).fetchone())
        return result


@app.post("/api/payments")
def add_payment(payload: PaymentIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid = user["business_id"]
    with db() as conn:
        party = conn.execute("SELECT * FROM parties WHERE id=? AND business_id=?", (payload.party_id, bid)).fetchone()
        if not party:
            raise HTTPException(status_code=404, detail="Party not found")
        new_balance = max(0.0, round(party["balance"] - payload.amount, 2))
        conn.execute("UPDATE parties SET balance=?,updated_at=? WHERE id=?", (new_balance, now_iso(), payload.party_id))
        entry_type = "payment_received" if payload.payment_type == "received" else "payment_paid"
        conn.execute(
            "INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (bid, payload.party_id, normalize_date(payload.payment_date), entry_type, "payment", 0, payload.amount, f"{payload.mode}: {payload.note}".strip(), now_iso()),
        )
        adjust_account(conn, bid, payload.mode, payload.amount if payload.payment_type == "received" else -payload.amount)
        updated = dict(conn.execute("SELECT * FROM parties WHERE id=?", (payload.party_id,)).fetchone())
    return updated


ALIASES: dict[str, list[str]] = {
    "name": ["item_name", "product_name", "party_name", "customer_name", "supplier_name", "name"],
    "sku": ["item_code", "product_code", "sku", "code"],
    "barcode": ["barcode", "bar_code"],
    "category": ["category", "item_category"],
    "unit": ["unit", "primary_unit"],
    "size": ["size", "item_size", "pack_size", "variant", "batch_size"],
    "hsn": ["hsn", "hsn_sac", "hsn_code"],
    "gst_rate": ["tax_rate", "gst", "gst_rate", "tax_percentage"],
    "purchase_price": ["purchase_price", "purchase_rate", "buy_price"],
    "sale_price": ["sale_price", "selling_price", "sale_rate"],
    "mrp": ["mrp", "maximum_retail_price"],
    "stock": ["stock", "opening_stock", "current_stock", "quantity"],
    "min_stock": ["min_stock", "low_stock_alert", "minimum_stock"],
    "type": ["party_type", "type"],
    "phone": ["phone", "mobile", "mobile_number", "contact_number"],
    "gstin": ["gstin", "gst_number", "gst_no"],
    "address": ["address", "billing_address"],
    "opening_balance": ["opening_balance", "balance", "current_balance"],
    "invoice_no": ["invoice_no", "invoice_number", "bill_no", "bill_number", "ref_no"],
    "invoice_date": ["invoice_date", "date", "bill_date"],
    "item_name": ["item_name", "product_name", "description"],
    "qty": ["qty", "quantity", "item_qty"],
    "rate": ["rate", "price", "item_rate"],
    "discount": ["discount", "discount_amount"],
    "paid": ["paid", "paid_amount", "received_amount"],
    "payment_mode": ["payment_mode", "payment_type", "mode"],
}


def pick(row: dict[str, Any], key: str, default: Any = "") -> Any:
    for alias in ALIASES.get(key, [key]):
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return default


def parse_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    rows: list[dict[str, Any]] = []
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        try:
            headers = [normalize_header(x) for x in next(values)]
        except StopIteration:
            return []
        for values_row in values:
            row = {headers[i]: values_row[i] for i in range(min(len(headers), len(values_row))) if headers[i]}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return rows
    if suffix == ".xls":
        if xlrd is None:
            raise ValueError("XLS support is not installed")
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return []
        headers = [normalize_header(ws.cell_value(0, c)) for c in range(ws.ncols)]
        for r in range(1, ws.nrows):
            row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols) if headers[c]}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return rows
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    for raw in reader:
        row = {normalize_header(k): v for k, v in raw.items() if k is not None}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def find_or_create_item(
    conn: sqlite3.Connection,
    bid: int,
    name: str,
    sku: str = "",
    size: str = "",
    unit: str = "pcs",
    batch_id: int | None = None,
) -> int:
    clean_name, clean_size = split_item_name_size(name, size, unit)
    row = None
    if sku:
        row = conn.execute("SELECT id FROM items WHERE business_id=? AND sku=?", (bid, sku)).fetchone()
    if not row:
        row = conn.execute(
            "SELECT id FROM items WHERE business_id=? AND lower(name)=lower(?) AND lower(COALESCE(size,''))=lower(?) LIMIT 1",
            (bid, clean_name, clean_size),
        ).fetchone()
    if row:
        return int(row["id"])
    generated_sku = sku or f"IMP-{secrets.token_hex(3).upper()}"
    cur = conn.execute(
        "INSERT INTO items(business_id,name,sku,unit,size,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (bid, clean_name or "Imported Item", generated_sku, unit or "pcs", clean_size, now_iso(), now_iso()),
    )
    return int(cur.lastrowid)


def find_or_create_party(conn: sqlite3.Connection, bid: int, name: str, party_type: str) -> int | None:
    if not name or name.lower() in {"cash", "cash customer", "cash supplier"}:
        return None
    row = conn.execute("SELECT id,type FROM parties WHERE business_id=? AND lower(name)=lower(?)", (bid, name)).fetchone()
    if row:
        if row["type"] != party_type and row["type"] != "both":
            conn.execute("UPDATE parties SET type='both',updated_at=? WHERE id=?", (now_iso(), row["id"]))
        return int(row["id"])
    cur = conn.execute(
        "INSERT INTO parties(business_id,name,type,created_at,updated_at) VALUES(?,?,?,?,?)",
        (bid, name, party_type, now_iso(), now_iso()),
    )
    return int(cur.lastrowid)


@app.post("/api/import/vyapar")
async def import_vyapar(
    entity_type: Literal["items", "parties", "sales", "purchases"] = Form(...),
    dry_run: bool = Form(False),
    file: UploadFile = File(...),
    user: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    content = await file.read()
    try:
        rows = parse_upload(file.filename or "import.csv", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Import file could not be read: {exc}") from exc
    preview = rows[:10]
    if dry_run:
        return {"dry_run": True, "entity_type": entity_type, "rows_total": len(rows), "headers": list(rows[0].keys()) if rows else [], "preview": preview}
    bid = user["business_id"]
    imported = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO import_batches(business_id,entity_type,filename,status,rows_total,created_at) VALUES(?,?,?,?,?,?)",
            (bid, entity_type, file.filename or "import", "processing", len(rows), now_iso()),
        )
        batch_id = int(cur.lastrowid)
        if entity_type == "items":
            for index, row in enumerate(rows, start=2):
                try:
                    raw_name = str(pick(row, "name", "")).strip()
                    if not raw_name:
                        raise ValueError("Item name missing")
                    raw_unit = str(pick(row, "unit", "pcs")).strip() or "pcs"
                    raw_size = str(pick(row, "size", "")).strip()
                    name, item_size = split_item_name_size(raw_name, raw_size, raw_unit)
                    sku = str(pick(row, "sku", "")).strip() or f"IMP-{index:05d}"
                    existing = conn.execute(
                        "SELECT id FROM items WHERE business_id=? AND (sku=? OR (lower(name)=lower(?) AND lower(COALESCE(size,''))=lower(?))) LIMIT 1",
                        (bid, sku, name, item_size),
                    ).fetchone()
                    values = (
                        name, sku, str(pick(row, "barcode", "")).strip(), str(pick(row, "category", "")).strip(), raw_unit, item_size, str(pick(row, "hsn", "")).strip(),
                        number(pick(row, "gst_rate", 0)), money(pick(row, "purchase_price", 0)), money(pick(row, "sale_price", 0)), money(pick(row, "mrp", 0)), number(pick(row, "stock", 0)), number(pick(row, "min_stock", 0)), now_iso(),
                    )
                    if existing:
                        conn.execute(
                            "UPDATE items SET name=?,sku=?,barcode=?,category=?,unit=?,size=?,hsn=?,gst_rate=?,purchase_price=?,sale_price=?,mrp=?,stock=?,min_stock=?,updated_at=? WHERE id=?",
                            values + (existing["id"],),
                        )
                    else:
                        conn.execute(
                            "INSERT INTO items(business_id,name,sku,barcode,category,unit,size,hsn,gst_rate,purchase_price,sale_price,mrp,stock,min_stock,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (bid,) + values[:-1] + (now_iso(), now_iso()),
                        )
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append({"row": index, "error": str(exc)})
        elif entity_type == "parties":
            for index, row in enumerate(rows, start=2):
                try:
                    name = str(pick(row, "name", "")).strip()
                    if not name:
                        raise ValueError("Party name missing")
                    raw_type = str(pick(row, "type", "customer")).strip().lower()
                    party_type = "supplier" if "supplier" in raw_type or "vendor" in raw_type else "customer"
                    existing = conn.execute("SELECT id FROM parties WHERE business_id=? AND lower(name)=lower(?)", (bid, name)).fetchone()
                    opening = money(pick(row, "opening_balance", 0))
                    values = (name, party_type, str(pick(row, "phone", "")).strip(), str(pick(row, "gstin", "")).strip(), str(pick(row, "address", "")).strip(), opening, opening, now_iso())
                    if existing:
                        conn.execute("UPDATE parties SET name=?,type=?,phone=?,gstin=?,address=?,opening_balance=?,balance=?,updated_at=? WHERE id=?", values + (existing["id"],))
                    else:
                        conn.execute("INSERT INTO parties(business_id,name,type,phone,gstin,address,opening_balance,balance,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)", (bid,) + values[:-1] + (now_iso(), now_iso()))
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append({"row": index, "error": str(exc)})
        else:
            grouped: dict[str, list[dict[str, Any]]] = {}
            for index, row in enumerate(rows, start=2):
                invoice_no = str(pick(row, "invoice_no", "")).strip() or f"IMP-{entity_type[:1].upper()}-{index:06d}"
                grouped.setdefault(invoice_no, []).append(row)
            for invoice_no, invoice_rows in grouped.items():
                try:
                    first = invoice_rows[0]
                    party_name = str(pick(first, "name", "")).strip()
                    party_type = "customer" if entity_type == "sales" else "supplier"
                    party_id = find_or_create_party(conn, bid, party_name, party_type)
                    tx_lines: list[TxLineIn] = []
                    for row in invoice_rows:
                        raw_item_name = str(pick(row, "item_name", pick(row, "name", "Imported Item"))).strip() or "Imported Item"
                        raw_line_size = str(pick(row, "size", "")).strip()
                        raw_line_unit = str(pick(row, "unit", "pcs")).strip() or "pcs"
                        item_name, line_size = split_item_name_size(raw_item_name, raw_line_size, raw_line_unit)
                        sku = str(pick(row, "sku", "")).strip()
                        item_id = find_or_create_item(conn, bid, item_name, sku, line_size, raw_line_unit, batch_id)
                        item = conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
                        tx_lines.append(TxLineIn(
                            item_id=item_id,
                            item_name=item_name,
                            size=line_size or str(item["size"] if item else "").strip(),
                            qty=max(number(pick(row, "qty", 1)), 0.0001),
                            rate=money(pick(row, "rate", item["sale_price"] if entity_type == "sales" else item["purchase_price"])),
                            gst_rate=number(pick(row, "gst_rate", item["gst_rate"])),
                        ))
                    payload = TransactionIn(
                        invoice_no=invoice_no,
                        party_id=party_id,
                        invoice_date=normalize_date(pick(first, "invoice_date", today_iso())),
                        discount=money(pick(first, "discount", 0)),
                        paid=money(pick(first, "paid", 0)),
                        payment_mode=str(pick(first, "payment_mode", "cash")),
                        notes="Imported from Vyapar",
                        items=tx_lines,
                    )
                    if entity_type == "sales":
                        insert_sale(conn, bid, payload, batch_id)
                    else:
                        insert_purchase(conn, bid, payload, batch_id)
                    imported += len(invoice_rows)
                except Exception as exc:
                    skipped += len(invoice_rows)
                    errors.append({"invoice_no": invoice_no, "error": str(exc)})
        conn.execute(
            "UPDATE import_batches SET status=?,rows_imported=?,rows_skipped=?,errors_json=? WHERE id=?",
            ("completed" if not errors else "completed_with_errors", imported, skipped, json.dumps(errors[:200], ensure_ascii=False), batch_id),
        )
    return {"batch_id": batch_id, "entity_type": entity_type, "rows_total": len(rows), "rows_imported": imported, "rows_skipped": skipped, "errors": errors[:50]}


@app.get("/api/import/batches")
def import_batches(user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    with db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM import_batches WHERE business_id=? ORDER BY id DESC LIMIT 100", (user["business_id"],)).fetchall()]
    for row in rows:
        row["errors"] = json.loads(row.pop("errors_json") or "[]")
    return rows


@app.get("/api/accounts")
def list_accounts(user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    with db() as conn:
        if conn.execute("SELECT COUNT(*) FROM accounts WHERE business_id=?", (user["business_id"],)).fetchone()[0] == 0:
            conn.execute("INSERT INTO accounts(business_id,name,account_type,balance,is_default,created_at,updated_at) VALUES(?,?,?,?,?,?,?)", (user["business_id"], "Cash In Hand", "cash", 0, 1, now_iso(), now_iso()))
        return [dict(r) for r in conn.execute("SELECT * FROM accounts WHERE business_id=? ORDER BY is_default DESC,name", (user["business_id"],)).fetchall()]


@app.post("/api/accounts")
def create_account(payload: AccountIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    with db() as conn:
        cur = conn.execute("INSERT INTO accounts(business_id,name,account_type,balance,is_default,created_at,updated_at) VALUES(?,?,?,?,0,?,?)", (user["business_id"], payload.name.strip(), payload.account_type, payload.opening_balance, now_iso(), now_iso()))
        return dict(conn.execute("SELECT * FROM accounts WHERE id=?", (cur.lastrowid,)).fetchone())


@app.get("/api/entries")
def list_entries(entry_type: str = "", limit: int = Query(default=200, ge=1, le=1000), user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    sql="SELECT * FROM business_entries WHERE business_id=?"
    args:[Any]=[user["business_id"]]
    if entry_type:
        sql += " AND entry_type=?"
        args.append(entry_type)
    sql += " ORDER BY entry_date DESC,id DESC LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(r) for r in conn.execute(sql,args).fetchall()]


@app.post("/api/entries")
def create_entry(payload: EntryIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid=user["business_id"]
    with db() as conn:
        party_name=""
        if payload.party_id:
            party=conn.execute("SELECT name FROM parties WHERE id=? AND business_id=?",(payload.party_id,bid)).fetchone()
            if not party: raise HTTPException(status_code=404, detail="Party not found")
            party_name=party["name"]
        account_id=payload.account_id
        if not account_id:
            row=conn.execute("SELECT id FROM accounts WHERE business_id=? ORDER BY is_default DESC,id LIMIT 1",(bid,)).fetchone()
            account_id=row["id"] if row else None
        # Balance direction: money-in adds, money-out subtracts, transfer moves between accounts.
        outgoing={"expense","cash_out","cheque_paid","loan_out","asset_purchase"}
        incoming={"cash_in","cheque_received","loan_in","asset_sale"}
        if payload.entry_type == "transfer":
            if not account_id or not payload.to_account_id or account_id == payload.to_account_id:
                raise HTTPException(status_code=400, detail="Transfer needs two different accounts")
            conn.execute("UPDATE accounts SET balance=balance-?,updated_at=? WHERE id=? AND business_id=?",(payload.amount,now_iso(),account_id,bid))
            conn.execute("UPDATE accounts SET balance=balance+?,updated_at=? WHERE id=? AND business_id=?",(payload.amount,now_iso(),payload.to_account_id,bid))
        elif payload.entry_type in outgoing and account_id:
            conn.execute("UPDATE accounts SET balance=balance-?,updated_at=? WHERE id=? AND business_id=?",(payload.amount,now_iso(),account_id,bid))
        elif payload.entry_type in incoming and account_id:
            conn.execute("UPDATE accounts SET balance=balance+?,updated_at=? WHERE id=? AND business_id=?",(payload.amount,now_iso(),account_id,bid))
        if payload.party_id and payload.entry_type in {"cash_in", "cash_out"}:
            current = conn.execute("SELECT balance FROM parties WHERE id=? AND business_id=?", (payload.party_id, bid)).fetchone()
            new_balance = max(0.0, round(float(current["balance"]) - payload.amount, 2)) if current else 0
            conn.execute("UPDATE parties SET balance=?,updated_at=? WHERE id=?", (new_balance, now_iso(), payload.party_id))
            conn.execute("INSERT INTO ledger_entries(business_id,party_id,entry_date,entry_type,reference_type,debit,credit,note,created_at) VALUES(?,?,?,?,?,?,?,?,?)", (bid, payload.party_id, normalize_date(payload.entry_date), "payment_received" if payload.entry_type == "cash_in" else "payment_paid", "business_entry", 0, payload.amount, payload.note or payload.title, now_iso()))
        cur=conn.execute("INSERT INTO business_entries(business_id,entry_type,entry_date,title,party_id,party_name,account_id,to_account_id,amount,status,mode,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",(bid,payload.entry_type,normalize_date(payload.entry_date),payload.title.strip(),payload.party_id,party_name,account_id,payload.to_account_id,payload.amount,payload.status,payload.mode,payload.note,now_iso()))
        return dict(conn.execute("SELECT * FROM business_entries WHERE id=?",(cur.lastrowid,)).fetchone())


@app.get("/api/documents")
def list_documents(kind: str = "", limit: int = Query(default=200, ge=1, le=1000), user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    sql="SELECT * FROM documents WHERE business_id=?"
    args:[Any]=[user["business_id"]]
    if kind:
        sql += " AND kind=?"
        args.append(kind)
    sql += " ORDER BY doc_date DESC,id DESC LIMIT ?"
    args.append(limit)
    with db() as conn:
        return [dict(r) for r in conn.execute(sql,args).fetchall()]


def next_document_no(conn: sqlite3.Connection, bid: int, kind: str) -> str:
    prefix={"sale_return":"SR","purchase_return":"PR","delivery_challan":"DC","estimate":"EST","proforma":"PI","sale_order":"SO","purchase_order":"PO","sale_asset":"SA","purchase_asset":"PA"}.get(kind,"DOC")
    row=conn.execute("SELECT id FROM documents WHERE business_id=? AND kind=? ORDER BY id DESC LIMIT 1",(bid,kind)).fetchone()
    seq=(row["id"]+1) if row else 1
    return f"{prefix}-{date.today().year % 100:02d}-{seq:05d}"


@app.post("/api/documents")
def create_document(payload: DocumentIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    bid=user["business_id"]
    with db() as conn:
        party_name=""
        if payload.party_id:
            party=conn.execute("SELECT name FROM parties WHERE id=? AND business_id=?",(payload.party_id,bid)).fetchone()
            if not party: raise HTTPException(status_code=404, detail="Party not found")
            party_name=party["name"]
        doc_no=payload.doc_no.strip() or next_document_no(conn,bid,payload.kind)
        cur=conn.execute("INSERT INTO documents(business_id,kind,doc_no,doc_date,party_id,party_name,amount,status,note,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",(bid,payload.kind,doc_no,normalize_date(payload.doc_date),payload.party_id,party_name,payload.amount,payload.status,payload.note,now_iso()))
        return dict(conn.execute("SELECT * FROM documents WHERE id=?",(cur.lastrowid,)).fetchone())


@app.get("/api/activity")
def activity(limit: int = Query(default=50, ge=1, le=500), user: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    bid=user["business_id"]
    with db() as conn:
        sales=[dict(r) for r in conn.execute("SELECT id,invoice_no AS ref,party_name AS title,invoice_date AS entry_date,total AS amount,due,'sale' AS kind FROM sales WHERE business_id=? ORDER BY id DESC LIMIT ?",(bid,limit)).fetchall()]
        purchases=[dict(r) for r in conn.execute("SELECT id,invoice_no AS ref,party_name AS title,invoice_date AS entry_date,total AS amount,due,'purchase' AS kind FROM purchases WHERE business_id=? ORDER BY id DESC LIMIT ?",(bid,limit)).fetchall()]
        entries=[dict(r) for r in conn.execute("SELECT id,title AS ref,COALESCE(NULLIF(party_name,''),title) AS title,entry_date,amount,0 AS due,entry_type AS kind,status FROM business_entries WHERE business_id=? ORDER BY id DESC LIMIT ?",(bid,limit)).fetchall()]
        docs=[dict(r) for r in conn.execute("SELECT id,doc_no AS ref,COALESCE(NULLIF(party_name,''),kind) AS title,doc_date AS entry_date,amount,0 AS due,kind,status FROM documents WHERE business_id=? ORDER BY id DESC LIMIT ?",(bid,limit)).fetchall()]
        returns=[dict(r) for r in conn.execute("SELECT id,return_no AS ref,party_name AS title,return_date AS entry_date,total AS amount,due,kind,'completed' AS status FROM returns WHERE business_id=? ORDER BY id DESC LIMIT ?",(bid,limit)).fetchall()]
    return sorted(sales+purchases+entries+docs+returns,key=lambda x:(x.get("entry_date", ""),x.get("id",0)),reverse=True)[:limit]


@app.get("/api/reports/summary")
def report_summary(
    date_from: str = "",
    date_to: str = "",
    user: dict[str, Any] = Depends(current_user),
) -> dict[str, Any]:
    bid = user["business_id"]
    start = normalize_date(date_from) if date_from else date.today().replace(day=1).isoformat()
    end = normalize_date(date_to) if date_to else today_iso()
    with db() as conn:
        sales = conn.execute("SELECT COALESCE(SUM(total),0),COALESCE(SUM(tax),0),COALESCE(SUM(due),0),COUNT(*) FROM sales WHERE business_id=? AND invoice_date BETWEEN ? AND ?", (bid, start, end)).fetchone()
        purchases = conn.execute("SELECT COALESCE(SUM(total),0),COALESCE(SUM(tax),0),COALESCE(SUM(due),0),COUNT(*) FROM purchases WHERE business_id=? AND invoice_date BETWEEN ? AND ?", (bid, start, end)).fetchone()
        sale_returns = conn.execute("SELECT COALESCE(SUM(total),0),COALESCE(SUM(tax),0),COUNT(*) FROM returns WHERE business_id=? AND kind='sale_return' AND return_date BETWEEN ? AND ?", (bid, start, end)).fetchone()
        purchase_returns = conn.execute("SELECT COALESCE(SUM(total),0),COALESCE(SUM(tax),0),COUNT(*) FROM returns WHERE business_id=? AND kind='purchase_return' AND return_date BETWEEN ? AND ?", (bid, start, end)).fetchone()
        stock_value = conn.execute("SELECT COALESCE(SUM(stock*purchase_price),0) FROM items WHERE business_id=? AND COALESCE(archived_at,'')=''", (bid,)).fetchone()[0]
        top_items = [dict(r) for r in conn.execute(
            """
            SELECT si.item_name,COALESCE(si.size,'') AS size,SUM(si.qty) AS qty,SUM(si.line_total) AS amount
            FROM sale_items si JOIN sales s ON s.id=si.sale_id
            WHERE s.business_id=? AND s.invoice_date BETWEEN ? AND ?
            GROUP BY si.item_name,COALESCE(si.size,'') ORDER BY amount DESC LIMIT 10
            """, (bid, start, end)
        ).fetchall()]
    return {
        "date_from": start,
        "date_to": end,
        "sales": {"amount": sales[0], "tax": sales[1], "due": sales[2], "count": sales[3]},
        "purchases": {"amount": purchases[0], "tax": purchases[1], "due": purchases[2], "count": purchases[3]},
        "sale_returns": {"amount": sale_returns[0], "tax": sale_returns[1], "count": sale_returns[2]},
        "purchase_returns": {"amount": purchase_returns[0], "tax": purchase_returns[1], "count": purchase_returns[2]},
        "net_sales": round(sales[0] - sale_returns[0], 2),
        "net_purchases": round(purchases[0] - purchase_returns[0], 2),
        "gross_margin_estimate": round((sales[0] - sale_returns[0]) - (purchases[0] - purchase_returns[0]), 2),
        "stock_value": round(stock_value, 2),
        "top_items": top_items,
    }


@app.get("/api/export/items.csv")
def export_items(user: dict[str, Any] = Depends(current_user)):
    from fastapi.responses import StreamingResponse
    with db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT name,sku,barcode,category,unit,size,hsn,gst_rate,purchase_price,sale_price,mrp,stock,min_stock FROM items WHERE business_id=? AND COALESCE(archived_at,'')='' ORDER BY name", (user["business_id"],)).fetchall()]
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=kirana-items.csv"})


def csv_response(filename: str, headers: list[str], rows: Iterable[Iterable[Any]]):
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/export/parties.csv")
def export_parties(user: dict[str, Any] = Depends(current_user)):
    with db() as conn:
        rows = conn.execute("SELECT name,type,phone,gstin,address,opening_balance,balance FROM parties WHERE business_id=? ORDER BY name", (user["business_id"],)).fetchall()
    return csv_response("kirana-parties.csv", ["Name","Type","Phone","GSTIN","Address","Opening Balance","Current Balance"], ([r[k] for k in r.keys()] for r in rows))


@app.get("/api/export/sales.csv")
def export_sales(user: dict[str, Any] = Depends(current_user)):
    with db() as conn:
        rows = conn.execute(
            """SELECT s.invoice_no,s.invoice_date,s.party_name,si.item_name,si.size,si.qty,si.rate,si.gst_rate,si.line_total,s.discount,s.paid,s.due,s.payment_mode
               FROM sales s JOIN sale_items si ON si.sale_id=s.id WHERE s.business_id=? ORDER BY s.invoice_date,s.id,si.id""",
            (user["business_id"],),
        ).fetchall()
    return csv_response("kirana-sales.csv", ["Invoice No","Date","Party","Item","Size","Qty","Rate","GST","Line Total","Bill Discount","Paid","Due","Mode"], ([r[k] for k in r.keys()] for r in rows))


@app.get("/api/export/purchases.csv")
def export_purchases(user: dict[str, Any] = Depends(current_user)):
    with db() as conn:
        rows = conn.execute(
            """SELECT p.invoice_no,p.invoice_date,p.party_name,pi.item_name,pi.size,pi.qty,pi.rate,pi.gst_rate,pi.line_total,p.discount,p.paid,p.due,p.payment_mode
               FROM purchases p JOIN purchase_items pi ON pi.purchase_id=p.id WHERE p.business_id=? ORDER BY p.invoice_date,p.id,pi.id""",
            (user["business_id"],),
        ).fetchall()
    return csv_response("kirana-purchases.csv", ["Invoice No","Date","Party","Item","Size","Qty","Rate","GST","Line Total","Bill Discount","Paid","Due","Mode"], ([r[k] for k in r.keys()] for r in rows))


@app.get("/api/backup/database")
def backup_database(user: dict[str, Any] = Depends(current_user)) -> FileResponse:
    if not DB_PATH.exists():
        raise HTTPException(status_code=404, detail="Database not found")
    with db() as conn:
        conn.execute("PRAGMA wal_checkpoint(FULL)")
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return FileResponse(DB_PATH, media_type="application/octet-stream", filename=f"kirana-backup-{stamp}.db")


if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=STATIC_DIR), name="assets")


@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/{path:path}")
def spa_fallback(path: str):
    candidate = STATIC_DIR / path
    if candidate.exists() and candidate.is_file():
        return FileResponse(candidate)
    return FileResponse(STATIC_DIR / "index.html")
