from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from fastapi import Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import load_workbook

import backend.app as core
import backend.import_fix_ext as import_fix
import backend.sale_import_ext  # noqa: F401 - load the earlier safety cleanup first
from backend.app import app, current_user, db, now_iso


PREVIOUS_PARSE = core.parse_upload

for key, aliases in {
    "invoice_no": ["invoice_no_txn_no"],
    "rate": ["unitprice", "unit_price"],
    "paid": ["received_paid_amount"],
    "phone": ["phone_no", "party_phone_no"],
}.items():
    for alias in aliases:
        if alias not in core.ALIASES.setdefault(key, []):
            core.ALIASES[key].append(alias)


def _compact_filename(filename: str) -> str:
    return re.sub(r"[^a-z]", "", Path(filename or "").stem.lower())


def _id(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\.0+$", "", text) if re.fullmatch(r"\d+\.0+", text) else text


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _rows(sheet: Any) -> list[dict[str, Any]]:
    # Vyapar places a Generated-on row before the actual header. The existing
    # smart detector scans the first 50 rows and finds the real header.
    return import_fix._rows_from_matrix(list(sheet.iter_rows(values_only=True)))


def _party_rows(workbook: Any) -> list[dict[str, Any]]:
    sheet = workbook["Party Report"] if "Party Report" in workbook.sheetnames else workbook.active
    output: list[dict[str, Any]] = []
    for row in _rows(sheet):
        name = str(row.get("name") or "").strip()
        if not name or _clean(name) == "total":
            continue
        receivable = core.money(row.get("receivable_balance"), 0)
        payable = core.money(row.get("payable_balance"), 0)
        supplier = payable > 0 and receivable <= 0
        output.append({
            **row,
            "name": name,
            "type": "supplier" if supplier else "customer",
            "phone": str(row.get("phone_no") or "").strip(),
            "opening_balance": payable if supplier else receivable,
        })
    return output


def _summary_rows(workbook: Any, kind: str) -> list[dict[str, Any]]:
    title = "Sale Report" if kind == "sales" else "Purchase Report"
    sheet = workbook[title] if title in workbook.sheetnames else workbook.active
    wanted = "sale" if kind == "sales" else "purchase"
    output: list[dict[str, Any]] = []
    for index, row in enumerate(_rows(sheet), start=1):
        tx_type = _clean(row.get("transaction_type"))
        if tx_type and tx_type != wanted:
            continue
        date_value = row.get("date")
        party = str(row.get("party_name") or "").strip()
        if not date_value or not party:
            continue
        output.append({
            "invoice_no": _id(row.get("invoice_no")),
            "invoice_date": core.normalize_date(date_value),
            "name": party,
            "total": core.money(row.get("total_amount"), 0),
            "paid": core.money(row.get("received_paid_amount"), 0),
            "payment_mode": str(row.get("payment_type") or "cash").strip() or "cash",
            "source_index": index,
        })
    return output


def _item_rows(workbook: Any, kind: str) -> list[dict[str, Any]]:
    if "Item Details" not in workbook.sheetnames:
        raise ValueError("Item Details sheet nahi mili")
    wanted = "sale" if kind == "sales" else "purchase"
    output: list[dict[str, Any]] = []
    for index, row in enumerate(_rows(workbook["Item Details"]), start=1):
        tx_type = _clean(row.get("transaction_type"))
        if tx_type and tx_type != wanted:
            continue
        item_name = str(row.get("item_name") or "").strip()
        party = str(row.get("party_name") or "").strip()
        date_value = row.get("date")
        qty = core.number(row.get("quantity"), 0)
        if not item_name or not party or not date_value or qty <= 0:
            continue
        gst = core.number(row.get("tax_percent"), 0)
        amount = core.money(row.get("amount"), 0)
        tax_amount = core.money(row.get("tax"), 0)
        unit_price = core.money(row.get("unitprice"), 0)
        if amount:
            before_tax = amount - tax_amount if tax_amount else amount / (1 + gst / 100) if gst else amount
            rate = max(0.0, round(before_tax / qty, 6))
        else:
            rate = unit_price
            amount = round(qty * rate * (1 + gst / 100), 2)
        output.append({
            **row,
            "invoice_no": _id(row.get("invoice_no_txn_no")),
            "invoice_date": core.normalize_date(date_value),
            "name": party,
            "item_name": item_name,
            "sku": _id(row.get("item_code")),
            "size": str(row.get("size") or "").strip(),
            "unit": str(row.get("unit") or "pcs").strip() or "pcs",
            "qty": qty,
            "rate": rate,
            "gst_rate": gst,
            "line_total": amount,
            # Line discount is already included in Amount/rate.
            "discount": 0,
            "source_index": index,
        })
    return output


def _partition(amounts: list[float], targets: list[float]) -> list[tuple[int, int]]:
    """Split ordered blank-invoice item rows against ordered summary totals."""
    n, m = len(amounts), len(targets)
    if not n or not m:
        return []
    prefix = [0.0]
    for value in amounts:
        prefix.append(prefix[-1] + value)
    inf = float("inf")
    dp = [[inf] * (n + 1) for _ in range(m + 1)]
    prev: list[list[int | None]] = [[None] * (n + 1) for _ in range(m + 1)]
    dp[0][0] = 0.0
    for group in range(1, m + 1):
        target = targets[group - 1]
        for end in range(n + 1):
            for start in range(end + 1):
                if dp[group - 1][start] == inf:
                    continue
                total = prefix[end] - prefix[start]
                cost = abs(total - target) / max(1.0, abs(target))
                if start == end and n >= m:
                    cost += 3.0
                value = dp[group - 1][start] + cost
                if value < dp[group][end]:
                    dp[group][end] = value
                    prev[group][end] = start
    if prev[m][n] is None:
        return [(0, n)]
    cuts: list[tuple[int, int]] = []
    end = n
    for group in range(m, 0, -1):
        start = prev[group][end]
        if start is None:
            return [(0, n)]
        cuts.append((start, end))
        end = start
    return list(reversed(cuts))


def _apply_summary(row: dict[str, Any], summary: dict[str, Any], invoice_no: str) -> None:
    row["invoice_no"] = invoice_no
    row["invoice_date"] = summary["invoice_date"]
    row["name"] = summary["name"]
    row["paid"] = summary["paid"]
    row["payment_mode"] = summary["payment_mode"]
    row["expected_total"] = summary["total"]


def _merge_summary(items: list[dict[str, Any]], summaries: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
    exact = {(row["invoice_date"], row["invoice_no"]): row for row in summaries if row["invoice_no"]}
    by_invoice: dict[str, list[dict[str, Any]]] = defaultdict(list)
    blank_summary: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in summaries:
        if row["invoice_no"]:
            by_invoice[row["invoice_no"]].append(row)
        else:
            blank_summary[(row["invoice_date"], _clean(row["name"]))].append(row)

    blank_items: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in items:
        invoice = row["invoice_no"]
        if invoice:
            summary = exact.get((row["invoice_date"], invoice))
            if summary is None and len(by_invoice.get(invoice, [])) == 1:
                summary = by_invoice[invoice][0]
            if summary:
                _apply_summary(row, summary, invoice)
        else:
            blank_items[(row["invoice_date"], _clean(row["name"]))].append(row)

    prefix = "S" if kind == "sales" else "P"
    for key, rows in blank_items.items():
        summaries_for_key = blank_summary.get(key, [])
        if summaries_for_key:
            cuts = _partition(
                [core.money(row["line_total"], 0) for row in rows],
                [core.money(row["total"], 0) for row in summaries_for_key],
            )
            for summary, (start, end) in zip(summaries_for_key, cuts):
                if start == end:
                    continue
                invoice = f"VYP-{prefix}-{summary['invoice_date'].replace('-', '')}-{summary['source_index']:06d}"
                for row in rows[start:end]:
                    _apply_summary(row, summary, invoice)
        for row in rows:
            if not row.get("invoice_no"):
                row["invoice_no"] = f"VYP-{prefix}-{row['invoice_date'].replace('-', '')}-{rows[0]['source_index']:06d}"

    # Invoice numbers restart in older Vyapar financial years. Add the date only
    # where the same number occurs on multiple dates inside one uploaded file.
    invoice_dates: dict[str, set[str]] = defaultdict(set)
    for row in items:
        invoice_dates[row["invoice_no"]].add(row["invoice_date"])
    repeated = {invoice for invoice, dates in invoice_dates.items() if len(dates) > 1}
    for row in items:
        if row["invoice_no"] in repeated:
            row["invoice_no"] = f"{row['invoice_no']}-{row['invoice_date'].replace('-', '')}"

    # Make line calculations equal Vyapar's summary total (round-off/additional
    # charge differences are placed into the last line rate).
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in items:
        grouped[row["invoice_no"]].append(row)
    for rows in grouped.values():
        expected = next((core.money(row.get("expected_total"), 0) for row in rows if row.get("expected_total") not in (None, "")), None)
        if expected is None:
            continue
        calculated = sum(row["qty"] * row["rate"] * (1 + row["gst_rate"] / 100) for row in rows)
        difference = expected - calculated
        if abs(difference) > 0.009:
            last = rows[-1]
            factor = 1 + last["gst_rate"] / 100
            last["rate"] = max(0.0, round(last["rate"] + difference / (last["qty"] * factor), 6))
    return items


def parse_exact_vyapar(filename: str, content: bytes) -> list[dict[str, Any]]:
    compact = _compact_filename(filename)
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return PREVIOUS_PARSE(filename, content)
    if not any(name in compact for name in ("partyreport", "salereport", "purchasereport")):
        return PREVIOUS_PARSE(filename, content)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if "partyreport" in compact:
            return _party_rows(workbook)
        kind = "sales" if "salereport" in compact else "purchases"
        items = _item_rows(workbook, kind)
        if not items:
            raise ValueError("Item Details sheet mein usable item data nahi mila")
        summaries = _summary_rows(workbook, kind)
        return _merge_summary(items, summaries, kind)
    finally:
        workbook.close()


core.parse_upload = parse_exact_vyapar


def _wrong_purchase_batches(conn: Any, business_id: int) -> list[dict[str, Any]]:
    batches = conn.execute(
        "SELECT * FROM import_batches WHERE business_id=? AND entity_type='purchases' AND COALESCE(status,'')!='rolled_back' ORDER BY id DESC",
        (business_id,),
    ).fetchall()
    result = []
    for batch in batches:
        purchases = conn.execute(
            "SELECT id,party_name FROM purchases WHERE business_id=? AND import_batch_id=?",
            (business_id, batch["id"]),
        ).fetchall()
        if not purchases:
            continue
        bad = True
        for purchase in purchases:
            lines = conn.execute("SELECT item_name FROM purchase_items WHERE purchase_id=?", (purchase["id"],)).fetchall()
            if len(lines) != 1 or _clean(lines[0]["item_name"]) not in {
                _clean(purchase["party_name"]), "imported item", "imported sale",
            }:
                bad = False
                break
        if bad:
            result.append({"batch_id": int(batch["id"]), "filename": batch["filename"], "transactions": len(purchases)})
    return result


def _rollback_wrong_purchase_batch(conn: Any, business_id: int, batch_id: int) -> None:
    purchases = conn.execute(
        "SELECT id,party_id,due,paid,payment_mode FROM purchases WHERE business_id=? AND import_batch_id=?",
        (business_id, batch_id),
    ).fetchall()
    for purchase in purchases:
        for line in conn.execute("SELECT item_id,qty FROM purchase_items WHERE purchase_id=?", (purchase["id"],)).fetchall():
            if line["item_id"]:
                conn.execute("UPDATE items SET stock=stock-?,updated_at=? WHERE id=?", (float(line["qty"] or 0), now_iso(), line["item_id"]))
        conn.execute("DELETE FROM stock_movements WHERE business_id=? AND reference_type='purchase' AND reference_id=?", (business_id, purchase["id"]))
        conn.execute("DELETE FROM ledger_entries WHERE business_id=? AND reference_type='purchase' AND reference_id=?", (business_id, purchase["id"]))
        if purchase["party_id"] and purchase["due"]:
            conn.execute("UPDATE parties SET balance=MAX(0,balance-?),updated_at=? WHERE id=?", (float(purchase["due"] or 0), now_iso(), purchase["party_id"]))
        if purchase["paid"]:
            core.adjust_account(conn, business_id, purchase["payment_mode"], float(purchase["paid"] or 0))
    conn.execute("DELETE FROM purchases WHERE business_id=? AND import_batch_id=?", (business_id, batch_id))
    conn.execute(
        "UPDATE import_batches SET status='rolled_back',rows_imported=0,errors_json=? WHERE id=? AND business_id=?",
        (json.dumps([{"error": "Removed summary-only purchase import"}]), batch_id, business_id),
    )
    conn.execute(
        """
        DELETE FROM items WHERE business_id=? AND sku LIKE 'IMP-%'
          AND ABS(COALESCE(stock,0))<0.000001
          AND NOT EXISTS (SELECT 1 FROM sale_items WHERE sale_items.item_id=items.id)
          AND NOT EXISTS (SELECT 1 FROM purchase_items WHERE purchase_items.item_id=items.id)
          AND NOT EXISTS (SELECT 1 FROM return_items WHERE return_items.item_id=items.id)
        """,
        (business_id,),
    )


@app.post("/api/import/cleanup-wrong-purchases")
def cleanup_wrong_purchases(execute: bool = Query(False), user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner access required")
    with db() as conn:
        batches = _wrong_purchase_batches(conn, user["business_id"])
        if execute:
            for batch in batches:
                _rollback_wrong_purchase_batch(conn, user["business_id"], batch["batch_id"])
    return {
        "execute": execute,
        "batch_count": len(batches),
        "transaction_count": sum(batch["transactions"] for batch in batches),
        "batches": batches,
    }


V4_JS = r"""
(() => {
  'use strict';
  loadImportHistory = async function () {
    const [rows, wrongParties, wrongSales, wrongPurchases] = await Promise.all([
      api('/api/import/batches'),
      api('/api/import/cleanup-party-items', {method:'POST'}),
      api('/api/import/cleanup-empty-sales', {method:'POST'}),
      api('/api/import/cleanup-wrong-purchases', {method:'POST'}),
    ]);
    let warning = '';
    if (wrongParties.count) warning += `<div class="info-banner" style="margin-bottom:12px"><b>${wrongParties.count} party names Items mein mile</b><p>Zero-stock aur unused wrong items hi remove honge.</p><button id="cleanup-wrong-party-items" class="btn primary" type="button">Remove Wrong Party Items</button></div>`;
    if (wrongSales.transaction_count) warning += `<div class="info-banner" style="margin-bottom:12px;border-color:#ef476f"><b>${wrongSales.transaction_count} blank ₹0 sales galat import hui</b><p>Sirf blank Imported Sale bills remove honge.</p><button id="cleanup-empty-sales" class="btn primary" type="button">Remove Blank Wrong Sales</button></div>`;
    if (wrongPurchases.transaction_count) warning += `<div class="info-banner" style="margin-bottom:12px;border-color:#ef476f"><b>${wrongPurchases.transaction_count} purchases summary sheet se galat import hui</b><p>Galat stock/items remove karke PurchaseReport dobara upload karein.</p><button id="cleanup-wrong-purchases" class="btn primary" type="button">Remove Wrong Purchases</button></div>`;
    const history = rows.map(row => { const error=row.errors?.[0]?.error||''; const removed=row.status==='rolled_back'?' · Removed':''; return `<div class="simple-row"><div><b>${esc(row.entity_type)} · ${esc(row.filename)}</b><small>${niceDate(row.created_at?.slice(0,10))}${removed}${error?` · ${esc(error)}`:''}</small></div><strong>${row.rows_imported}/${row.rows_total}</strong></div>`; }).join('') || emptyText('Abhi koi import nahi hua.');
    $('#import-history').innerHTML = warning + history;
  };
  document.addEventListener('click', async event => {
    const saleButton = event.target.closest('#cleanup-empty-sales');
    if (saleButton) {
      if (!confirm('Sirf blank ₹0 Imported Sale bills remove karein?')) return;
      saleButton.disabled = true;
      try { const result=await api('/api/import/cleanup-empty-sales?execute=true',{method:'POST'}); toast(`${result.transaction_count} wrong sales removed`); await refreshAll(); await loadImportHistory(); }
      catch(error){ toast(error.message,true); saleButton.disabled=false; }
      return;
    }
    const purchaseButton = event.target.closest('#cleanup-wrong-purchases');
    if (purchaseButton) {
      if (!confirm('Summary sheet se galat import hui purchases aur unka stock remove karein?')) return;
      purchaseButton.disabled = true;
      try { const result=await api('/api/import/cleanup-wrong-purchases?execute=true',{method:'POST'}); toast(`${result.transaction_count} wrong purchases removed`); await refreshAll(); await loadImportHistory(); }
      catch(error){ toast(error.message,true); purchaseButton.disabled=false; }
    }
  });
})();
"""


for route in list(app.router.routes):
    if getattr(route, "path", None) == "/import-fix.js":
        app.router.routes.remove(route)


@app.get("/import-fix.js", include_in_schema=False)
def import_fix_js_v4() -> Response:
    return Response(
        f"{import_fix.IMPORT_FIX_JS}\n{V4_JS}",
        media_type="application/javascript",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


paths = {"/api/import/cleanup-wrong-purchases", "/import-fix.js"}
routes = [route for route in app.router.routes if getattr(route, "path", None) in paths]
for route in routes:
    app.router.routes.remove(route)
fallback_index = next(
    (index for index, route in enumerate(app.router.routes) if getattr(route, "path", None) == "/{path:path}"),
    len(app.router.routes),
)
app.router.routes[fallback_index:fallback_index] = routes
